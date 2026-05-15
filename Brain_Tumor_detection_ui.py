
import sys
import os
import threading
from datetime import datetime
from PySide6.QtWidgets import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from pathlib import Path
from ultralytics import YOLO
from utils.TumorSliceFinder import TumorSliceFinder
import nibabel as nib
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# 处理打包环境
if getattr(sys, 'frozen', False):
    # 打包环境
    base_dir = Path(sys._MEIPASS)
else:
    # 开发环境
    base_dir = Path(__file__).parent

def get_resource_path(relative_path):
    """获取资源文件的绝对路径（兼容打包环境）"""
    if getattr(sys, 'frozen', False):
        # 打包后的环境
        base_path = Path(sys._MEIPASS)
    else:
        # 开发环境
        base_path = Path(__file__).parent
    return base_path / relative_path


def get_runtime_dir(relative_path):
    """获取运行时目录的绝对路径（用于创建历史记录等运行时数据）
    
    注意：与 get_resource_path 不同，运行时目录应该在可执行文件所在目录创建，
    而不是在打包的临时目录中。
    """
    if getattr(sys, 'frozen', False):
        # 打包后的环境：在可执行文件所在目录创建
        base_path = Path(sys.executable).parent
    else:
        # 开发环境：在项目根目录创建
        base_path = Path(__file__).parent
    return base_path / relative_path


def get_runtime_dir(relative_path):
    """获取运行时目录的绝对路径（用于创建历史记录等运行时数据）"""
    if getattr(sys, 'frozen', False):
        # 打包后的环境：在可执行文件所在目录创建
        base_path = Path(sys.executable).parent
    else:
        # 开发环境：在项目根目录创建
        base_path = Path(__file__).parent
    return base_path / relative_path


def get_runtime_dir(relative_path):
    """获取运行时目录的绝对路径（用于创建历史记录等运行时数据）"""
    if getattr(sys, 'frozen', False):
        # 打包后的环境：在可执行文件所在目录创建
        base_path = Path(sys.executable).parent
    else:
        # 开发环境：在项目根目录创建
        base_path = Path(__file__).parent
    return base_path / relative_path

class StyleManager:
    """Style Manager - Provides gradient and modern UI styles"""


    @staticmethod
    def get_main_stylesheet():
        return """
            QMainWindow {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8f9fa, stop:1 #e9ecef);
            }

            QGroupBox {
                font-weight: bold;
                font-size: 12px;
                border: 2px solid rgba(52, 152, 219, 0.7);
                border-radius: 8px;
                margin-top: 1ex;
                padding-top: 15px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.9), stop:1 rgba(245, 245, 245, 0.9));
            }

            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 10px 0 10px;
                color: #2c3e50;
                font-size: 13px;
                font-weight: bold;
            }

            QPushButton {
                padding: 2px 8px;
                font-size: 12px;
                font-weight: bold;
                border: none;
                border-radius: 8px;
                color: white;
                min-width: 65px;
                min-height: 25px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
            }

            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #5dade2, stop:1 #3498db);
                transform: translateY(-1px);
            }

            QPushButton:pressed {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2980b9, stop:1 #1f618d);
            }

            QPushButton:disabled {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #bdc3c7, stop:1 #95a5a6);
                color: #7f8c8d;
            }

            QComboBox {
                padding: 2px 8px;
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 8px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 white, stop:1 #f8f9fa);
                font-size: 12px;
                min-width: 150px;
                min-height: 25px;
            }

            QComboBox:focus {
                border-color: #3498db;
                background: white;
            }

            QProgressBar {
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 8px;
                text-align: center;
                font-weight: bold;
                font-size: 11px;
                max-height: 20px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ecf0f1, stop:1 #d5dbdb);
            }

            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2ecc71, stop:1 #27ae60);
                border-radius: 6px;
                margin: 1px;
            }

            QTextEdit {
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 8px;
                background: rgba(255, 255, 255, 0.95);
                font-family: 'Consolas', 'Monaco', monospace;
                font-size: 11px;
                padding: 8px;
                selection-background-color: #3498db;
            }

            QSlider::groove:horizontal {
                border: 1px solid rgba(189, 195, 199, 0.5);
                height: 8px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ecf0f1, stop:1 #bdc3c7);
                border-radius: 4px;
            }

            QSlider::handle:horizontal {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                border: 2px solid #2980b9;
                width: 20px;
                height: 20px;
                margin: -8px 0;
                border-radius: 12px;
            }

            QSlider::handle:horizontal:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #5dade2, stop:1 #3498db);
            }

            QSpinBox, QDoubleSpinBox {
                padding: 6px 10px;
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 6px;
                background: white;
                min-width: 80px;
                font-size: 12px;
            }

            QTabWidget::pane {
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 10px;
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(255, 255, 255, 0.95), stop:1 rgba(245, 245, 245, 0.95));
                margin-top: 5px;
            }

            QTabBar::tab {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #ecf0f1, stop:1 #bdc3c7);
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-bottom: none;
                border-radius: 8px 8px 0 0;
                padding: 12px 25px;
                margin-right: 3px;
                font-weight: bold;
                font-size: 12px;
                color: #2c3e50;
            }

            QTabBar::tab:selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                color: white;
                border-color: rgba(52, 152, 219, 0.7);
            }

            QTabBar::tab:hover:!selected {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #d5dbdb, stop:1 #bdc3c7);
            }

            QTableWidget {
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 8px;
                background: white;
                gridline-color: rgba(189, 195, 199, 0.3);
                selection-background-color: rgba(52, 152, 219, 0.2);
                alternate-background-color: rgba(248, 249, 250, 0.5);
            }

            QTableWidget::item {
                padding: 8px;
                border: none;
            }

            QTableWidget::item:selected {
                background: rgba(52, 152, 219, 0.3);
                color: #2c3e50;
            }

            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #34495e, stop:1 #2c3e50);
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }

            QListWidget {
                border: 2px solid rgba(189, 195, 199, 0.5);
                border-radius: 8px;
                background: white;
                selection-background-color: rgba(52, 152, 219, 0.2);
            }

            QListWidget::item {
                padding: 8px;
                border-bottom: 1px solid rgba(189, 195, 199, 0.2);
            }

            QListWidget::item:selected {
                background: rgba(52, 152, 219, 0.3);
                color: #2c3e50;
            }

            QScrollBar:vertical {
                background: rgba(236, 240, 241, 0.5);
                width: 12px;
                border-radius: 6px;
            }

            QScrollBar::handle:vertical {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #bdc3c7, stop:1 #95a5a6);
                border-radius: 6px;
                min-height: 20px;
            }

            QScrollBar::handle:vertical:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #95a5a6, stop:1 #7f8c8d);
            }
            #startBtn {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #10B981, stop:1 #059669);
            color: white;
            border: none;
            border-radius: 3px;  /* 减小圆角 */
            min-width: 10px;  /* 减小最小宽度 */
            min-height: 10px;  /* 添加最小高度 */
            font-size: 8px;  /* 减小字体大小 */
        }
        #startBtn:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #059669, stop:1 #047857);
        }

        #pauseBtn {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #F59E0B, stop:1 #D97706);
            color: white;
            border: none;
            border-radius: 3px;  /* 减小圆角 */
            min-width: 10px;  /* 减小最小宽度 */
            min-height: 10px;  /* 添加最小高度 */
            font-size: 8px;  /* 减小字体大小 */
        }
        #pauseBtn:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #D97706, stop:1 #B45309);
        }

        #stopBtn {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #EF4444, stop:1 #DC2626);
            color: white;
            border: none;
            border-radius: 3px;  /* 减小圆角 */
            min-width: 10px;  /* 减小最小宽度 */
            min-height: 10px;  /* 添加最小高度 */
            font-size: 8px;  /* 减小字体大小 */
        }
        #stopBtn:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #DC2626, stop:1 #B91C1C);
        }

        #monitorBtn {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #3B82F6, stop:1 #2563EB);
            color: white;
            border: none;
            border-radius: 3px;  /* 减小圆角 */
            min-width: 10px;  /* 减小最小宽度 */
            min-height: 10px;  /* 添加最小高度 */
            font-size: 8px;  /* 减小字体大小 */
        }
        #monitorBtn:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #2563EB, stop:1 #1D4ED8);
        }

        #clearBtn {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #94A3B8, stop:1 #64748B);
            color: white;
            border: none;
            border-radius: 3px;  /* 减小圆角 */
            min-width: 10px;  /* 减小最小宽度 */
            min-height: 10px;  /* 添加最小高度 */
            font-size: 8px;  /* 减小字体大小 */
        }
        #clearBtn:hover {
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                       stop:0 #64748B, stop:1 #475569);
        }
        """

    @staticmethod
    def get_image_label_style():
        return """
            border: 3px solid rgba(52, 152, 219, 0.3);
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                stop:0 rgba(248, 249, 250, 0.9), stop:1 rgba(233, 236, 239, 0.9));
            color: #7f8c8d;
            font-weight: bold;
            font-size: 12px;
            border-radius: 10px;
            padding: 15px;
        """
    @staticmethod
    def get_image_niigz_label_style():
        return """
            background: black;
            color: white;
            font-weight: bold;
            font-size: 12px;
            padding: 0px;
        """


    @staticmethod
    def get_video_label_style():
        return """
            border: 1px solid rgba(52, 152, 219, 0.3);
            background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                stop:0 rgba(248, 249, 250, 0.9), stop:1 rgba(233, 236, 239, 0.9));
            color: #7f8c8d;
            font-weight: bold;
            font-size: 14px;
            border-radius: 10px;
        """


class CameraManager:
    """Camera Manager - Handles multi-camera detection and management"""


    def __init__(self):
        self.cameras = []
        self.scan_cameras()

    def scan_cameras(self):
        """Scan available cameras"""
        self.cameras = []

        # Suppress OpenCV warnings during camera detection
        import sys
        import os
        
        # Save original stderr
        original_stderr = sys.stderr
        
        try:
            # Redirect stderr to suppress OpenCV warnings
            sys.stderr = open(os.devnull, 'w')
            
            # Detect cameras (check first 2 indices - most systems have 0-1 cameras)
            MAX_CAMERAS = 2
            
            for i in range(MAX_CAMERAS):
                try:
                    # Explicitly specify CAP_DSHOW backend for Windows stability
                    cap = cv2.VideoCapture(i, cv2.CAP_DSHOW)
                    
                    if not cap.isOpened():
                        cap.release()
                        continue
                    
                    # Use grab() instead of read() for safer availability check
                    if not cap.grab():
                        cap.release()
                        continue
                    
                    ret, frame = cap.read()
                    if not ret or frame is None:
                        cap.release()
                        continue
                    
                    # Get camera information
                    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
                    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
                    fps = cap.get(cv2.CAP_PROP_FPS)

                    camera_info = {
                        'id': i,
                        'name': f"Camera {i}",
                        'resolution': f"{width}x{height}",
                        'fps': fps if fps > 0 else 30,
                        'available': True,
                        'cap': None  # Reserve camera object position
                    }
                    self.cameras.append(camera_info)
                    
                    # Properly release camera
                    cap.release()
                    del cap
                    
                    # Found a working camera, but continue to check for more
                    
                except Exception as e:
                    print(f"Error detecting camera {i}: {e}")
                    try:
                        cap.release()
                    except:
                        pass
                    continue
        finally:
            # Restore original stderr
            sys.stderr.close()
            sys.stderr = original_stderr

        # If no cameras, add virtual camera for testing
        if not self.cameras:
            self.cameras.append({
                'id': -1,
                'name': "Virtual Camera",
                'resolution': "640x480",
                'fps': 30,
                'available': False,
                'cap': None
            })

    def get_camera_count(self):
        """Return available camera count

        Returns:
            int: Available camera count
        """
        return len(self.get_available_cameras())

    def get_available_cameras(self):
        """Get available camera list

        Returns:
            list: Available camera information dictionary list
        """
        return [cam for cam in self.cameras if cam['available']]

    def get_camera_info(self, camera_id):
        """Get detailed information for specified camera

        Args:
            camera_id (int): Camera ID

        Returns:
            dict: Camera information dictionary with following keys:
                - id: Camera index
                - name: Camera name
                - resolution: Resolution string (e.g., "640x480")
                - fps: Frame rate
                - available: Whether available
        """
        for cam in self.cameras:
            if cam['id'] == camera_id:
                return cam
        return None

    def get_camera_names(self):
        """Get all camera names list

        Returns:
            list: Camera name string list
        """
        return [cam['name'] for cam in self.cameras]

    def release_all(self):
        """Release all camera resources"""
        for cam in self.cameras:
            if cam['cap'] is not None:
                cam['cap'].release()
                cam['cap'] = None


class ModelManager:
    """Model Manager - Handles model scanning and loading"""

    def __init__(self):
        self.models_paths = [
            get_resource_path("pt_models"),
            get_resource_path("onnx_models"),
            get_resource_path("../models"),
            get_resource_path("weights"),
        ]
        self.current_model = None
        self.class_names = []

    def scan_models(self, custom_path=None):
        """Scan model files"""
        models = []
        search_paths = self.models_paths.copy()

        if custom_path and Path(custom_path).exists():
            search_paths.insert(0, Path(custom_path))

        for model_dir in search_paths:
            if model_dir.exists():
                try:
                    # Scan .pt files
                    pt_files = sorted(model_dir.glob("*.pt"))
                    for pt_file in pt_files:
                        models.append({
                            'name': pt_file.name,
                            'path': str(pt_file),
                            'size': self._get_file_size(pt_file),
                            'modified': self._get_modification_time(pt_file),
                            'format': 'pt'
                        })
                    
                    # Scan .onnx files
                    onnx_files = sorted(model_dir.glob("*.onnx"))
                    for onnx_file in onnx_files:
                        models.append({
                            'name': onnx_file.name,
                            'path': str(onnx_file),
                            'size': self._get_file_size(onnx_file),
                            'modified': self._get_modification_time(onnx_file),
                            'format': 'onnx'
                        })
                except Exception as e:
                    print(f"Error scanning directory {model_dir}: {e}")

        return models

    def load_model(self, model_path):
        """Load model"""
        try:
            # Check file extension to determine model format
            file_extension = Path(model_path).suffix.lower()
            
            if file_extension == '.onnx':
                # Load ONNX model
                self.current_model = YOLO(model_path, task='detect')
            else:
                # Default load .pt model
                self.current_model = YOLO(model_path)
            
            self.class_names = list(self.current_model.names.values())
            return True
        except Exception as e:
            print(f"Model loading failed: {e}")
            return False

    def get_class_names(self):
        """Get class names"""
        return self.class_names

    def _get_file_size(self, file_path):
        """Get file size"""
        try:
            size = file_path.stat().st_size
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size < 1024.0:
                    return f"{size:.1f} {unit}"
                size /= 1024.0
            return f"{size:.1f} TB"
        except:
            return "Unknown"

    def _get_modification_time(self, file_path):
        """Get modification time"""
        try:
            timestamp = file_path.stat().st_mtime
            return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M")
        except:
            return "Unknown"


class DetectionThread(QThread):
    """增强的检测线程"""
    result_ready = Signal(object, object, float, object, list)  # 原图, 结果图, 耗时, 检测结果, 类别名称
    progress_updated = Signal(int)
    status_changed = Signal(str)
    error_occurred = Signal(str)
    fps_updated = Signal(float)
    finished = Signal()

    def __init__(self, model, source_type, source_path=None, camera_id=0, confidence_threshold=0.25):
        super().__init__()
        self.model = model
        self.source_type = source_type
        self.source_path = source_path
        self.camera_id = camera_id
        self.confidence_threshold = confidence_threshold
        self.is_running = False
        self.is_paused = False
        self.frame_count = 0
        self.fps_counter = 0
        self.last_fps_time = time.time()

    def run(self):
        self.is_running = True
        try:
            if self.source_type == 'image':
                self._process_image()
            elif self.source_type == 'video':
                self._process_video()
            elif self.source_type == 'camera':
                self._process_camera()
        except Exception as e:
            self.error_occurred.emit(f"Detection error: {str(e)}")
        finally:
            self.is_running = False
            self.finished.emit()

    def _process_image(self):
        """处理单张图片"""
        if not self.source_path or not Path(self.source_path).exists():
            self.error_occurred.emit("Image file not found")
            return

        self.status_changed.emit("Processing image...")

        start_time = time.time()
        results = self.model(self.source_path, conf=self.confidence_threshold, verbose=False)
        end_time = time.time()

        original_img = cv2.imread(self.source_path)
        if original_img is None:
            self.error_occurred.emit("Image file read error")
            return

        original_img = cv2.cvtColor(original_img, cv2.COLOR_BGR2RGB)
        result_img = results[0].plot()
        result_img = cv2.cvtColor(result_img, cv2.COLOR_BGR2RGB)
        class_names = list(self.model.names.values())

        self.result_ready.emit(original_img, result_img, end_time - start_time, results, class_names)
        self.progress_updated.emit(100)
        self.status_changed.emit("Detection completed")

    def _process_video(self):
        """处理视频文件"""
        if not self.source_path or not Path(self.source_path).exists():
            self.error_occurred.emit("Video file not found")
            return

        # Use CAP_FFMPEG for video files (more compatible)
        cap = cv2.VideoCapture(self.source_path, cv2.CAP_FFMPEG)
        if not cap.isOpened():
            self.error_occurred.emit("Video file open error")
            return

        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        frame_count = 0
        class_names = list(self.model.names.values())

        self.status_changed.emit(f"Processing video... {total_frames} frames)")

        while cap.isOpened() and self.is_running:
            if self.is_paused:
                time.sleep(0.1)
                continue

            ret, frame = cap.read()
            if not ret:
                break

            start_time = time.time()
            results = self.model(frame, conf=self.confidence_threshold, verbose=False)
            end_time = time.time()

            original_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            result_img = results[0].plot()
            result_img = cv2.cvtColor(result_img, cv2.COLOR_BGR2RGB)

            self.result_ready.emit(original_img, result_img, end_time - start_time, results, class_names)

            frame_count += 1
            if total_frames > 0:
                progress = int((frame_count / total_frames) * 100)
                self.progress_updated.emit(progress)

            # 更新FPS
            self._update_fps()

            # 状态更新（每30帧更新一次）
            if frame_count % 30 == 0:
                current_fps = self._get_current_fps()
                self.status_changed.emit(f"Processing video... {frame_count}/{total_frames} frames (FPS: {current_fps:.1f})")

            time.sleep(0.033)  # 约30fps

        cap.release()

    def _process_camera(self):
        """处理摄像头"""
        # Explicitly specify CAP_DSHOW backend for Windows stability
        cap = cv2.VideoCapture(self.camera_id, cv2.CAP_DSHOW)
        if not cap.isOpened():
            self.error_occurred.emit(f"Camera {self.camera_id} open error")
            return

        # 设置摄像头参数
        cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
        cap.set(cv2.CAP_PROP_FPS, 30)

        class_names = list(self.model.names.values())
        self.status_changed.emit(f"Camera {self.camera_id} started...")

        while cap.isOpened() and self.is_running:
            if self.is_paused:
                time.sleep(0.1)
                continue

            ret, frame = cap.read()
            if not ret:
                break

            start_time = time.time()
            results = self.model(frame, conf=self.confidence_threshold, verbose=False)
            end_time = time.time()

            original_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            result_img = results[0].plot()
            result_img = cv2.cvtColor(result_img, cv2.COLOR_BGR2RGB)

            self.result_ready.emit(original_img, result_img, end_time - start_time, results, class_names)

            # 更新FPS
            self._update_fps()

            # 状态更新（每60帧更新一次）
            if self.frame_count % 60 == 0:
                current_fps = self._get_current_fps()
                self.status_changed.emit(f"Camera {self.camera_id} processing (FPS: {current_fps:.1f})")

            time.sleep(0.033)  # 约30fps

        cap.release()

    def _update_fps(self):
        """更新FPS计算"""
        self.frame_count += 1
        self.fps_counter += 1

        current_time = time.time()
        if current_time - self.last_fps_time >= 1.0:
            fps = self.fps_counter / (current_time - self.last_fps_time)
            self.fps_updated.emit(fps)
            self.fps_counter = 0
            self.last_fps_time = current_time

    def _get_current_fps(self):
        """获取当前FPS"""
        current_time = time.time()
        if current_time - self.last_fps_time > 0:
            return self.fps_counter / (current_time - self.last_fps_time)
        return 0

    def pause(self):
        self.is_paused = True
        self.status_changed.emit(f"Detection paused...")

    def resume(self):
        self.is_paused = False
        self.status_changed.emit(f"Detection resumed")

    def stop(self):
        self.is_running = False
        self.status_changed.emit(f"Detection completed!")


class BatchDetectionThread(QThread):
    """批量检测线程"""
    result_ready = Signal(str, object, object, float, object, list)  # 文件路径, 原图, 结果图, 耗时, 检测结果, 类别名称
    progress_updated = Signal(int)
    current_file_changed = Signal(str)
    status_changed = Signal(str)
    error_occurred = Signal(str)
    finished = Signal()

    def __init__(self, model, folder_path, confidence_threshold=0.25, supported_formats=None):
        super().__init__()
        self.model = model
        self.folder_path = folder_path
        self.confidence_threshold = confidence_threshold
        self.supported_formats = supported_formats or ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.webp', '.tif']
        self.is_running = False
        self.processed_count = 0
        self.error_count = 0

    def run(self):
        self.is_running = True

        try:
            # 收集所有支持的图片文件
            image_files = []
            for fmt in self.supported_formats:
                image_files.extend(Path(self.folder_path).rglob(f'*{fmt}'))
                # image_files.extend(Path(self.folder_path).rglob(f'*{fmt.upper()}'))

            total_files = len(image_files)
            if total_files == 0:
                self.status_changed.emit("No supported image formats found in folder")
                self.finished.emit()
                return

            self.status_changed.emit(f"Start batch processing {total_files} files...")

            # 获取类别名称
            class_names = list(self.model.names.values())

            for i, img_path in enumerate(image_files):
                if not self.is_running:
                    break

                self.current_file_changed.emit(str(img_path))

                try:
                    # 处理单个图片
                    start_time = time.time()
                    results = self.model(str(img_path), conf=self.confidence_threshold, verbose=False)
                    end_time = time.time()

                    # 获取原图
                    original_img = cv2.imread(str(img_path))
                    if original_img is not None:
                        original_img = cv2.cvtColor(original_img, cv2.COLOR_BGR2RGB)

                        # 获取结果图
                        result_img = results[0].plot()
                        result_img = cv2.cvtColor(result_img, cv2.COLOR_BGR2RGB)

                        self.result_ready.emit(str(img_path), original_img, result_img,
                                               end_time - start_time, results, class_names)
                        self.processed_count += 1

                except Exception as e:
                    self.error_occurred.emit(f"Processing file {img_path.name} error: {str(e)}")
                    self.error_count += 1

                # 更新进度
                progress = int(((i + 1) / total_files) * 100)
                self.progress_updated.emit(progress)

                # 状态更新
                if (i + 1) % 10 == 0 or i == total_files - 1:
                    self.status_changed.emit(
                        f"Batch processing progress: {i + 1}/{total_files} (Success: {self.processed_count}, Error: {self.error_count})")

        except Exception as e:
            self.error_occurred.emit(f"Batch processing error: {str(e)}")
        finally:
            self.is_running = False
            # self.finished.emit()

    def stop(self):
        """停止批量检测"""
        self.is_running = False


class MultiCameraMonitorThread(QThread):
    camera_result_ready = Signal(int, object, object, float, object, list)
    camera_error = Signal(int, str)
    camera_status = Signal(int, str)
    finished = Signal()

    def __init__(self, model, camera_ids, conf=0.25, fps=10):
        super().__init__()
        self.model = model
        self.cam_ids = camera_ids
        self.conf = conf
        self.period = 1.0 / fps  # Frame interval
        self.caps = {}  # {id: cv2.VideoCapture}
        self.active = {}  # {id: bool} Whether online
        self.last_t = {}  # {id: float}

        # Thread synchronization
        self._run_flag = True
        self._pause_cond = QWaitCondition()
        self._pause_mutex = QMutex()
        self._paused_flag = False

    # ----------------- Lifecycle -----------------
    def run(self):
        self._open_all()
        if not self.caps:
            self.finished.emit()
            return

        cls_names = list(self.model.names.values())

        while self._run_flag:
            self._pause_mutex.lock()
            if self._paused_flag:
                self._pause_cond.wait(self._pause_mutex)
            self._pause_mutex.unlock()

            for cid in list(self.caps.keys()):
                if not self._run_flag:
                    break
                if not self._grab_and_infer(cid, cls_names):
                    self._reconnect_later(cid)  # Async reconnect after disconnection
            self.msleep(10)

        self._close_all()
        self.finished.emit()

    def stop(self):
        self._run_flag = False
        self.resume()  # Ensure waiting thread is woken up
        self.wait()

    def pause(self):
        self._pause_mutex.lock()
        self._paused_flag = True
        self._pause_mutex.unlock()

    def resume(self):
        self._pause_mutex.lock()
        self._paused_flag = False
        self._pause_mutex.unlock()
        self._pause_cond.wakeAll()

    # ----------------- Private tools -----------------
    def _open_all(self):
        for cid in self.cam_ids:
            cap = cv2.VideoCapture(cid, cv2.CAP_DSHOW)
            if cap.isOpened():
                cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
                cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
                cap.set(cv2.CAP_PROP_FPS, 30)
                self.caps[cid] = cap
                self.active[cid] = True
                self.last_t[cid] = 0.0
                self.camera_status.emit(cid, "Connected")
            else:
                self.camera_error.emit(cid, "Cannot open")
                cap.release()

    def _close_all(self):
        for cap in self.caps.values():
            cap.release()
        self.caps.clear()

    def _grab_and_infer(self, cid, cls_names):
        cap = self.caps.get(cid)
        if not cap or not cap.isOpened():
            return False

        # Non-blocking frame reading: first grab then retrieve
        if not cap.grab():
            return False

        now = time.time()
        if now - self.last_t[cid] < self.period:
            return True  # Not timed out, but frame has been grabbed to avoid accumulation
        self.last_t[cid] = now

        ret, frame = cap.retrieve()
        if not ret:
            return False

        try:
            t0 = time.time()
            results = self.model(frame, conf=self.conf, verbose=False)
            infer_ms = (time.time() - t0) * 1000
            out_img = results[0].plot()
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            rgb_out = cv2.cvtColor(out_img, cv2.COLOR_BGR2RGB)
            self.camera_result_ready.emit(cid, rgb_frame, rgb_out,
                                          infer_ms / 1000.0, results, cls_names)
            return True
        except Exception as e:
            self.camera_error.emit(cid, f"Inference error: {e}")
            return False

    def _reconnect_later(self, cid):
        # Simple strategy: retry after 5 seconds
        if self.active.get(cid) is False:
            return
        self.active[cid] = False
        self.camera_status.emit(cid, "Reconnecting…")
        threading.Timer(5.0, lambda: self._try_reopen(cid)).start()

    def _try_reopen(self, cid):
        if cid in self.caps:
            self.caps[cid].release()
        # Explicitly specify CAP_DSHOW backend for reconnection
        cap = cv2.VideoCapture(cid, cv2.CAP_DSHOW)
        if cap.isOpened():
            self.caps[cid] = cap
            self.active[cid] = True
            self.camera_status.emit(cid, "Reconnected")
        else:
            cap.release()
            self._reconnect_later(cid)


class ModelSelectionDialog(QDialog):
    """Model selection dialog"""

    def __init__(self, model_manager, parent=None):
        super().__init__(parent)
        self.model_manager = model_manager
        self.selected_model = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("🔧 Advanced Model Selection")
        self.setModal(True)
        self.resize(700, 450)

        layout = QVBoxLayout(self)

        # Custom path
        path_group = QGroupBox("📁 Custom Model Path")
        path_layout = QHBoxLayout(path_group)

        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Enter custom model directory path...")
        path_layout.addWidget(self.path_edit)

        browse_btn = QPushButton("📂 Browse")
        browse_btn.clicked.connect(self.browse_path)
        path_layout.addWidget(browse_btn)

        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.clicked.connect(self.refresh_models)
        path_layout.addWidget(refresh_btn)

        layout.addWidget(path_group)

        # Model list
        models_group = QGroupBox("📋 Available Models")
        models_layout = QVBoxLayout(models_group)

        self.model_table = QTableWidget()
        self.model_table.setColumnCount(4)
        self.model_table.setHorizontalHeaderLabels(["Model Name", "Size", "Modified Time", "Path"])
        self.model_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.model_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.model_table.setAlternatingRowColors(True)
        self.model_table.doubleClicked.connect(self.accept)

        models_layout.addWidget(self.model_table)
        layout.addWidget(models_group)

        # Buttons
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

        # Set style
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #f8f9fa, stop:1 #e9ecef);
            }
        """)

        self.refresh_models()

    def browse_path(self):
        """Browse custom path"""
        path = QFileDialog.getExistingDirectory(self, "Select Model Directory")
        if path:
            self.path_edit.setText(path)
            self.refresh_models()

    def refresh_models(self):
        """Refresh model list"""
        custom_path = self.path_edit.text() if self.path_edit.text() else None
        models = self.model_manager.scan_models(custom_path)

        self.model_table.setRowCount(len(models))

        for i, model in enumerate(models):
            self.model_table.setItem(i, 0, QTableWidgetItem(model['name']))
            self.model_table.setItem(i, 1, QTableWidgetItem(model['size']))
            self.model_table.setItem(i, 2, QTableWidgetItem(model['modified']))
            self.model_table.setItem(i, 3, QTableWidgetItem(model['path']))

    def accept(self):
        """Confirm selection"""
        current_row = self.model_table.currentRow()
        if current_row >= 0:
            self.selected_model = self.model_table.item(current_row, 3).text()
        super().accept()



class DetectionResultWidget(QWidget):
    """Detection result display component - Modified version with diagnosis report"""

    def __init__(self):
        super().__init__()
        self.diagnosis_widget = None
        self.init_ui()

    def init_ui(self):
        # Use horizontal layout, left side detection result table, right side diagnosis report
        main_layout = QHBoxLayout(self)

        # Left panel: Detection result table area
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        # Title
        title = QLabel("🎯 Detection Result Details")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        left_layout.addWidget(title)

        # Result table
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels(
            ["Class", "Confidence", "Coordinates (x,y)", "Size (w×h)", "Estimated Diameter (pixel)", "Pixel Area (pixel²)"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.result_table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: white;  /* Deep blue background */
                color: #2c3e50;               
                font-size: 8pt;
                font-weight: bold;
                height: 12px;
                border: 1px solid #d0d0d0;       /* Border */
            }
        """)
        self.result_table.setStyleSheet("""
            QTableWidget {
                font-size: 8pt;  /* Adjust table content font size */
                text-align: center; /* Center alignment */
                align: center; /* Center alignment */
            }
        """)
        self.result_table.setMaximumHeight(200)
        self.result_table.setAlternatingRowColors(True)

        left_layout.addWidget(self.result_table)

        # Statistics information
        self.stats_label = QLabel("Detection summary...")
        self.stats_label.setMinimumHeight(60)  # Set minimum height to 40 pixels
        self.stats_label.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(236, 240, 241, 0.9), stop:1 rgba(189, 195, 199, 0.9));
            padding: 12px;
            border-radius: 8px;
            font-size: 12px;
            color: #2c3e50;
            font-weight: bold;
        """)
        # Modify this part of the code
        left_layout.addWidget(self.stats_label)

        main_layout.addWidget(left_panel)


    def update_results(self, results, class_names, inference_time):
        """Update detection results"""
        if not results or not results[0].boxes or len(results[0].boxes) == 0:
            self.result_table.setRowCount(0)
            self.stats_label.setText("❌ No targets detected")
            return

        boxes = results[0].boxes
        confidences = boxes.conf.cpu().numpy()
        classes = boxes.cls.cpu().numpy().astype(int)
        xyxy = boxes.xyxy.cpu().numpy()

        # Update table
        self.result_table.setRowCount(len(confidences))

        class_counts = {}
        detection_data = []  # Store detection data for report generation

        for i, (conf, cls, box) in enumerate(zip(confidences, classes, xyxy)):
            class_name = class_names[cls] if cls < len(class_names) else f"Class{cls}"

            # Count class occurrences
            class_counts[class_name] = class_counts.get(class_name, 0) + 1

            # Fill data starting from column 0, skipping serial number column
            self.result_table.setItem(i, 0, QTableWidgetItem(class_name))

            # Confidence with color
            conf_item = QTableWidgetItem(f"{conf:.3f}")
            if conf > 0.8:
                conf_item.setBackground(QColor(46, 204, 113, 100))  # Green
            elif conf > 0.5:
                conf_item.setBackground(QColor(241, 196, 15, 100))  # Yellow
            else:
                conf_item.setBackground(QColor(231, 76, 60, 100))  # Red
            self.result_table.setItem(i, 1, conf_item)

            # Calculate width and height
            width = box[2] - box[0]
            height = box[3] - box[1]

            self.result_table.setItem(i, 2, QTableWidgetItem(f"({box[0]:.0f},{box[1]:.0f})"))
            self.result_table.setItem(i, 3, QTableWidgetItem(f"{width:.0f}×{height:.0f}"))

            # 计算直径 (假设是圆形肿瘤，直径等于平均尺寸)
            diameter = (width + height) / 2  # 像素为单位
            diameter_mm = diameter  #
            self.result_table.setItem(i, 4, QTableWidgetItem(f"{diameter_mm:.2f}"))

            voxel_size = 1.0  # 默认体素单位为1mm
            voxel_value = 0.785 * width * height * voxel_size
            self.result_table.setItem(i, 5, QTableWidgetItem(f"{voxel_value:.2f}"))

            # 添加到检测数据列表，也相应调整
            detection_data.append({
                'index': i + 1,  # 仍然保留内部索引用于报告
                'class': class_name,
                'confidence': conf,
                'x': box[0],
                'y': box[1],
                'width': width,
                'height': height,
                'diameter': diameter_mm,
                'voxel': voxel_value
            })
        # 更新统计信息
        total_objects = len(confidences)
        if total_objects > 0:
            avg_confidence = np.mean(confidences)

            # 计算平均直径和体素
            diameters = []
            voxels = []

            for i, (conf, cls, box) in enumerate(zip(confidences, classes, xyxy)):
                width = box[2] - box[0]
                height = box[3] - box[1]

                # 计算直径 (假设是圆形肿瘤，直径等于平均尺寸)
                diameter = (width + height) / 2  # 像素为单位
                diameters.append(diameter)

                # 计算体素
                voxel_size = 1.0  # 默认体素单位为1mm
                voxel_value = 0.785 * width * height * voxel_size
                voxels.append(voxel_value)

            avg_diameter = np.mean(diameters)
            avg_voxel = np.mean(voxels)
            if total_objects > 1:
                # Build statistics text
                stats_text = f"✅ Detected {total_objects} objects | "
                stats_text += f"🎯 Average confidence: {avg_confidence:.3f} | "
                stats_text += f"⌀ Average diameter: {avg_diameter:.2f}pixel | "
                stats_text += f"🧩 Average pixel area: {avg_voxel:.2f}pixel²\n"
                stats_text += "📊 Class statistics: " + " | ".join([f"{name}: {count}" for name, count in class_counts.items()])
            else:
                # Build statistics text
                stats_text = f"✅ Detected {total_objects} object | "
                stats_text += f"🎯 Confidence: {avg_confidence:.3f} | "
                stats_text += f"⌀ Diameter: {avg_diameter:.2f}pixel | "
                stats_text += f"🧩 Pixel area: {avg_voxel:.2f}pixel²\n"
                stats_text += "📊 Class: " + " | ".join([f"{name}: {count}" for name, count in class_counts.items()])

        else:
            stats_text = "❌ No objects detected"

        self.stats_label.setText(stats_text)
class MonitoringWidget(QWidget):
    """Monitoring page component"""

    def __init__(self, model_manager, camera_manager):
        super().__init__()
        self.model_manager = model_manager
        self.camera_manager = camera_manager
        self.monitoring_thread = None
        self.camera_labels = {}
        self.current_model = None
        self.start_monitor_btn = QPushButton("🚀 Start monitoring")

        # Auto-save monitoring snapshots related properties
        self.is_auto_saving = False
        self.camera_recorders = {}  # {camera_id: VideoRecorder}
        self.monitor_history_dir = get_runtime_dir("monitor_history")
        self.monitor_history_dir.mkdir(exist_ok=True)
        self.current_memory_usage = 0  # MB
        self.max_memory_limit = 500  # MB

        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)

        # Control panel
        control_group = QGroupBox("🖥️ Monitoring control")
        control_group.setMaximumHeight(120)  # Increase height to accommodate new controls
        control_layout = QVBoxLayout(control_group)

        # First row: Model and camera selection
        first_row_layout = QHBoxLayout()

        # Model selection
        model_layout = QHBoxLayout()
        model_layout.addWidget(QLabel("Model:"))
        self.model_combo = QComboBox()
        self.model_combo.currentTextChanged.connect(self.on_model_changed)
        self.init_model_combo()
        model_layout.addWidget(self.model_combo)
        select_model_btn = QPushButton("🔧 Select model")
        select_model_btn.clicked.connect(self.select_model)
        model_layout.addWidget(select_model_btn)
        first_row_layout.addLayout(model_layout)

        # Camera selection
        camera_layout = QHBoxLayout()
        camera_layout.addWidget(QLabel("Camera:"))
        self.camera_list = QListWidget()
        self.camera_list.setMaximumWidth(300)
        self.camera_list.setSelectionMode(QListWidget.MultiSelection)
        self.refresh_cameras()
        camera_layout.addWidget(self.camera_list)
        refresh_camera_btn = QPushButton("🔄 Refresh")
        refresh_camera_btn.clicked.connect(self.refresh_cameras)
        camera_layout.addWidget(refresh_camera_btn)
        first_row_layout.addLayout(camera_layout)

        control_layout.addLayout(first_row_layout)

        # Second row: Monitoring control and auto-save settings
        second_row_layout = QHBoxLayout()

        # Monitoring control buttons
        monitor_btn_layout = QHBoxLayout()
        self.start_monitor_btn.clicked.connect(self.start_monitoring)
        self.start_monitor_btn.setEnabled(True)
        monitor_btn_layout.addWidget(self.start_monitor_btn)

        self.stop_monitor_btn = QPushButton("⏸️ Pause")
        self.stop_monitor_btn.clicked.connect(self.stop_monitoring)
        monitor_btn_layout.addWidget(self.stop_monitor_btn)

        self.clear_monitor_btn = QPushButton("🗑️ Clear Monitor")
        self.clear_monitor_btn.clicked.connect(self.clear_monitoring)
        self.clear_monitor_btn.setEnabled(False)
        self.stop_monitor_btn.setEnabled(False)
        monitor_btn_layout.addWidget(self.clear_monitor_btn)

        second_row_layout.addLayout(monitor_btn_layout)

        # Auto-save monitoring snapshots control
        snapshot_control_layout = QHBoxLayout()

        self.auto_save_btn = QPushButton("🎬 Auto-save Monitoring Snapshots")
        self.auto_save_btn.clicked.connect(self.toggle_auto_save)
        self.auto_save_btn.setEnabled(False)
        snapshot_control_layout.addWidget(self.auto_save_btn)

        # Recording settings
        snapshot_control_layout.addWidget(QLabel("FPS:"))
        self.recording_fps_spinbox = QSpinBox()
        self.recording_fps_spinbox.setRange(5, 60)
        self.recording_fps_spinbox.setValue(20)
        self.recording_fps_spinbox.setSuffix(" fps")
        snapshot_control_layout.addWidget(self.recording_fps_spinbox)

        snapshot_control_layout.addWidget(QLabel("Memory Limit:"))
        self.memory_limit_spinbox = QSpinBox()
        self.memory_limit_spinbox.setRange(100, 2000)
        self.memory_limit_spinbox.setValue(500)
        self.memory_limit_spinbox.setSuffix(" MB")
        snapshot_control_layout.addWidget(self.memory_limit_spinbox)

        second_row_layout.addLayout(snapshot_control_layout)

        control_layout.addLayout(second_row_layout)

        layout.addWidget(control_group)

        # 监控显示区域
        self.monitor_scroll = QScrollArea()
        self.monitor_scroll.setStyleSheet("""
            QScrollArea {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(236, 240, 241, 0.9),
                    stop:1 rgba(189, 195, 199, 0.9));
                border-radius: 8px;
            }
            QScrollArea > QWidget > QWidget {   /* viewport */
                background: transparent;
            }
            QScrollArea::corner {               /* 右下角空白三角 */
                background: transparent;
            }
        """)
        self.monitor_widget = QWidget()
        self.monitor_layout = QGridLayout(self.monitor_widget)
        self.monitor_scroll.setWidget(self.monitor_widget)
        self.monitor_scroll.setWidgetResizable(True)

        layout.addWidget(self.monitor_scroll)

    def init_model_combo(self):
        """Initialize model dropdown"""
        self.model_combo.clear()
        models = self.model_manager.scan_models()

        if not models:
            self.model_combo.addItem("No models available")
            self.model_combo.setEnabled(False)
        else:
            self.model_combo.addItems([model['name'] for model in models])
            self.model_combo.setEnabled(True)
            self.try_load_default_model()

    def try_load_default_model(self):
        """Try to load default model"""
        if self.model_combo.count() > 0 and self.model_combo.itemText(0) != "No models available":
            first_model = self.model_combo.itemText(0)
            self.load_model_by_name(first_model)

    def load_model_by_name(self, model_name):
        """Load model by name"""
        models = self.model_manager.scan_models()
        for model in models:
            if model['name'] == model_name:
                self.load_model(model['path'])
                break

    def on_model_changed(self, model_text):
        """Model selection changed"""
        if model_text != "No models available":
            self.load_model_by_name(model_text)

    def load_model(self, model_path):

        """Load model"""
        try:
            self.current_model = YOLO(model_path)
            self.start_monitor_btn.setEnabled(True)
            return True
        except Exception as e:
            pass

            return False

    def select_model(self):
        """Select model"""
        dialog = ModelSelectionDialog(self.model_manager, self)
        if dialog.exec() == QDialog.Accepted and dialog.selected_model:
            try:
                self.current_model = YOLO(dialog.selected_model)
                model_name = Path(dialog.selected_model).name
                self.model_combo.clear()
                self.model_combo.addItem(model_name)
                self.start_monitor_btn.setEnabled(True)
                QMessageBox.information(self, "Success", f"Model loaded successfully: {model_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Model loading failed: {str(e)}")

    def refresh_cameras(self):
        """Refresh camera list"""
        self.camera_manager.scan_cameras()
        self.camera_list.clear()

        for camera in self.camera_manager.get_available_cameras():
            item = QListWidgetItem(f"📹 {camera['name']} ({camera['resolution']})")
            item.setData(Qt.UserRole, camera['id'])
            self.camera_list.addItem(item)

    def start_monitoring(self):
        """Start Monitoring"""
        if not self.current_model:
            QMessageBox.warning(self, "Warning", "Please select a model first")
            return

        selected_items = self.camera_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", "Please select at least one camera")
            return

        camera_ids = [item.data(Qt.UserRole) for item in selected_items]

        # 清空之前的显示
        self.clear_monitor_display()
        self.clear_monitor_btn.setEnabled(True)

        # 创建显示标签
        self.create_camera_labels(camera_ids)
        # 设置等高宽
        self.set_equal_column_stretch()
        # 启动监控线程
        self.monitoring_thread = MultiCameraMonitorThread(self.current_model, camera_ids)
        self.monitoring_thread.camera_result_ready.connect(self.update_camera_display)
        self.monitoring_thread.camera_error.connect(self.handle_camera_error)
        self.monitoring_thread.finished.connect(self.on_monitoring_finished)

        self.monitoring_thread.start()

        self.start_monitor_btn.setEnabled(False)
        self.stop_monitor_btn.setEnabled(True)
        self.auto_save_btn.setEnabled(True)  # 启用自动保存按钮

    def stop_monitoring(self):
        """暂停/继续监控"""
        if self.monitoring_thread and self.monitoring_thread._run_flag:
            if self.monitoring_thread._paused_flag:  # 监测是否已暂停
                self.monitoring_thread.resume()  # 恢复
                self.stop_monitor_btn.setText("⏸️ 暂停")  # 按钮文字：暂停
            else:
                self.monitoring_thread.pause()  # 暂停
                self.stop_monitor_btn.setText("▶️ 继续")  # 按钮文字：继续

    def clear_monitoring(self):
        """停止监控"""
        self.monitoring_thread.stop()
        self.clear_monitor_display()

        # 停止自动保存
        if self.is_auto_saving:
            self.stop_auto_save()

        # 重置按钮状态
        self.start_monitor_btn.setEnabled(True)
        self.stop_monitor_btn.setEnabled(False)
        self.clear_monitor_btn.setEnabled(False)
        self.auto_save_btn.setEnabled(False)

    def create_camera_labels(self, camera_ids):
        """创建摄像头显示标签"""
        self.camera_labels = {}

        cols = 2  # 每行2个摄像头
        for i, camera_id in enumerate(camera_ids):
            row = i // cols
            col = i % cols

            # 创建摄像头组
            camera_group = QGroupBox(f"📹 Camera {camera_id}")
            camera_group.setStyleSheet("""
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(248, 249, 250, 0.9), stop:1 rgba(233, 236, 239, 0.9));
                color: #7f8c8d;
                font-weight: bold;
                font-size: 14px;
                border-radius: 10px;

            """)
            # camera_group.setMaximumHeight(350)
            camera_layout = QVBoxLayout(camera_group)
            self.start_btn = QPushButton("▶️")
            self.start_btn.setObjectName("startBtn")
            self.start_btn.setToolTip("Start detection")
            # self.start_btn.clicked.connect(self.start_detection)

            self.pause_btn = QPushButton("⏸️")
            self.pause_btn.setObjectName("pauseBtn")
            self.pause_btn.setToolTip("Pause detection")
            # self.pause_btn.clicked.connect(self.pause_detection)

            self.stop_btn = QPushButton("⏹️")
            self.stop_btn.setObjectName("stopBtn")
            self.stop_btn.setToolTip("Stop detection")
            # self.stop_btn.clicked.connect(self.stop_detection)

            self.monitor_btn = QPushButton("👁️")
            self.monitor_btn.setObjectName("monitorBtn")
            self.monitor_btn.setToolTip("Monitor mode switch")
            # self.monitor_btn.clicked.connect(self.toggle_monitor_mode)

            self.clear_btn = QPushButton("🗑️")
            self.clear_btn.setObjectName("clearBtn")
            self.clear_btn.setToolTip("Clear monitor")
            # self.clear_btn.clicked.connect(self.clear_frame)

            # 状态标签
            status_label = QLabel("Status: Initializing...")
            status_label.setStyleSheet("color: #7f8c8d; font-size: 10px;")

            control_layout = QHBoxLayout()
            control_layout.addWidget(self.start_btn)
            control_layout.addWidget(self.pause_btn)
            control_layout.addWidget(self.stop_btn)
            control_layout.addWidget(self.monitor_btn)
            control_layout.addWidget(self.clear_btn)
            control_layout.addStretch()
            control_layout.addWidget(status_label)

            # 图像显示标签
            image_label = QLabel("Waiting for connection...")
            image_label.setMinimumSize(300, 240)
            # image_label.setMaximumHeight(350)
            image_label.setStyleSheet("""
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(248, 249, 250, 0.9), stop:1 rgba(233, 236, 239, 0.9));
                color: #7f8c8d;
                font-weight: bold;
                font-size: 14px;
                border-radius: 10px;

            """)
            image_label.setAlignment(Qt.AlignCenter)
            image_label.setScaledContents(True)

            camera_layout.addWidget(image_label, stretch=6)

            camera_layout.addLayout(control_layout)
            camera_layout.addStretch()

            self.camera_labels[camera_id] = {
                'image': image_label,
                'status': status_label,
                'group': camera_group
            }
            self.setStyleSheet("""
                QPushButton#startBtn {
                    max-width: 24px;
                    text-align: left;
                    padding: 5px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    background-color: #f8f9fa;
                    color: #7f8c8d;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton#startBtn:hover {
                    background-color: #e9ecef;
                }
                QPushButton#startBtn:pressed {
                    background-color: #dcdcdc;
                }

                QPushButton#pauseBtn {
                    max-width: 24px;
                    text-align: left;
                    padding: 5px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    background-color: #f8f9fa;
                    color: #7f8c8d;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton#pauseBtn:hover {
                    background-color: #e9ecef;
                }
                QPushButton#pauseBtn:pressed {
                    background-color: #dcdcdc;
                }

                QPushButton#stopBtn {
                    max-width: 24px;
                    text-align: left;
                    padding: 5px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    background-color: #f8f9fa;
                    color: #7f8c8d;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton#stopBtn:hover {
                    background-color: #e9ecef;
                }
                QPushButton#stopBtn:pressed {
                    background-color: #dcdcdc;
                }

                QPushButton#monitorBtn {
                    max-width: 24px;
                    text-align: left;
                    padding: 5px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    background-color: #f8f9fa;
                    color: #7f8c8d;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton#monitorBtn:hover {
                    background-color: #e9ecef;
                }
                QPushButton#monitorBtn:pressed {
                    background-color: #dcdcdc;
                }

                QPushButton#clearBtn {
                    max-width: 24px;
                    text-align: left;
                    padding: 5px;
                    border: 1px solid #ccc;
                    border-radius: 5px;
                    background-color: #f8f9fa;
                    color: #7f8c8d;
                    font-weight: bold;
                    font-size: 14px;
                }
                QPushButton#clearBtn:hover {
                    background-color: #e9ecef;
                }
                QPushButton#clearBtn:pressed {
                    background-color: #dcdcdc;
                }

                QLabel {
                    color: #7f8c8d;
                    font-size: 10px;
                }
            """)
            self.monitor_layout.addWidget(camera_group, row, col)

    def set_equal_column_stretch(self):
        for c in range(self.monitor_layout.columnCount()):
            self.monitor_layout.setColumnStretch(c, 1)
        for r in range(self.monitor_layout.rowCount()):
            self.monitor_layout.setRowStretch(r, 1)

    def clear_monitor_display(self):
        """清空监控显示"""
        for camera_id in list(self.camera_labels.keys()):
            self.camera_labels[camera_id]['group'].deleteLater()
        self.camera_labels.clear()

    def update_camera_display(self, camera_id, original_img, result_img, inference_time, results, class_names):
        """Update camera display"""
        if camera_id not in self.camera_labels:
            return

        # Display result image
        self.display_image(result_img, self.camera_labels[camera_id]['image'])

        # Update status
        if results and results[0].boxes and len(results[0].boxes) > 0:
            object_count = len(results[0].boxes)
            self.camera_labels[camera_id]['status'].setText(
                f"Status: Detected {object_count} objects | Time: {inference_time:.3f}s"
            )

            # Add detection frame to auto-save system
            detection_info = {
                'results': results,
                'class_names': class_names,
                'inference_time': inference_time
            }
            self.add_detection_frame(camera_id, result_img, detection_info)
        else:
            self.camera_labels[camera_id]['status'].setText(
                f"Status: No objects | Time: {inference_time:.3f}s"
            )

    def handle_camera_error(self, camera_id, error_msg):
        """Handle camera error"""
        if camera_id in self.camera_labels:
            self.camera_labels[camera_id]['status'].setText(f"Error: {error_msg}")
            self.camera_labels[camera_id]['status'].setStyleSheet("color: red; font-size: 10px;")

    def on_monitoring_finished(self):
        """Monitoring finished"""
        self.start_monitor_btn.setEnabled(True)
        self.stop_monitor_btn.setEnabled(False)

        for camera_id in self.camera_labels:
            self.camera_labels[camera_id]['status'].setText("Status: Stopped")

    def display_image(self, img_array, label):
        """Display image"""
        if img_array is None:
            return

        height, width, channel = img_array.shape
        bytes_per_line = 3 * width
        q_image = QImage(img_array.data, width, height, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_image)
        scaled_pixmap = pixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        label.setPixmap(scaled_pixmap)

    def toggle_auto_save(self):
        """Toggle auto-save monitoring snapshots status"""
        if not self.is_auto_saving:
            self.start_auto_save()
        else:
            self.stop_auto_save()

    def start_auto_save(self):
        """Start auto-save monitoring snapshots"""
        if not self.current_model:
            QMessageBox.warning(self, "Warning", "Please select a model first")
            return

        self.is_auto_saving = True
        self.max_memory_limit = self.memory_limit_spinbox.value()

        self.auto_save_btn.setText("⏹️ Stop auto-save")

        # Disable setting controls
        self.recording_fps_spinbox.setEnabled(False)
        self.memory_limit_spinbox.setEnabled(False)

        QMessageBox.information(self, "Success", "Auto-save monitoring snapshots started")

    def stop_auto_save(self):
        """Stop auto-save monitoring snapshots"""
        self.is_auto_saving = False

        # Stop all recorders
        for recorder in self.camera_recorders.values():
            recorder.stop_recording()
        self.camera_recorders.clear()

        self.auto_save_btn.setText("🎬 Auto-save monitoring snapshots")
        # Enable setting controls
        self.recording_fps_spinbox.setEnabled(True)
        self.memory_limit_spinbox.setEnabled(True)

        QMessageBox.information(self, "Success", "Auto-save monitoring snapshots stopped")

    def add_detection_frame(self, camera_id, frame, detection_info):
        """Add detection frame to auto-save system"""
        if not self.is_auto_saving:
            return

        # Check if there are detection results
        if not detection_info or not detection_info.get('results'):
            return

        results = detection_info['results']
        if not hasattr(results[0], 'boxes') or not results[0].boxes or len(results[0].boxes) == 0:
            return

        # Get camera name
        camera_name = f"Camera{camera_id}"
        if camera_id in self.camera_labels:
            camera_name = f"Camera{camera_id}"

        # Create or get recorder
        if camera_id not in self.camera_recorders:
            self.camera_recorders[camera_id] = CameraVideoRecorder(
                camera_id, camera_name, self.monitor_history_dir,
                self.recording_fps_spinbox.value()
            )
            # Start recording
            self.camera_recorders[camera_id].start_recording()

        # Add frame to recorder
        self.camera_recorders[camera_id].add_frame(frame, detection_info)

        # Check memory usage
        self.check_memory_usage()

    def check_memory_usage(self):
        """Check memory usage, clean up oldest records when limit is exceeded"""
        # Calculate current memory usage
        total_size = 0
        for json_file in self.monitor_history_dir.glob("*.json"):
            mp4_file = json_file.with_suffix('.mp4')
            if mp4_file.exists():
                total_size += mp4_file.stat().st_size

        current_usage_mb = total_size / (1024 * 1024)

        if current_usage_mb > self.max_memory_limit:
            # Delete oldest records
            self.cleanup_oldest_records()

    def cleanup_oldest_records(self):
        """Clean up oldest records"""
        json_files = list(self.monitor_history_dir.glob("*.json"))
        if not json_files:
            return

        # Sort by modification time, delete oldest
        json_files.sort(key=lambda x: x.stat().st_mtime)

        for json_file in json_files[:len(json_files) // 4]:  # Delete 25% of oldest records
            mp4_file = json_file.with_suffix('.mp4')
            try:
                if json_file.exists():
                    json_file.unlink()
                if mp4_file.exists():
                    mp4_file.unlink()
            except Exception as e:
                print(f"Failed to clean up file {json_file}: {e}")


class CameraVideoRecorder:
    """摄像头视频录制器"""

    def __init__(self, camera_id, camera_name, output_dir, fps=20):
        self.camera_id = camera_id
        self.camera_name = camera_name
        self.output_dir = output_dir
        self.fps = fps
        self.is_recording = False
        self.video_writer = None
        self.frames = []
        self.detection_stats = {}
        self.total_detections = 0
        self.start_time = None
        self.end_time = None
        self.max_frames_per_file = fps * 30  # 30秒的视频

    def start_recording(self):
        """开始录制"""
        if self.is_recording:
            return

        self.is_recording = True
        self.start_time = time.time()
        self.frames.clear()
        self.detection_stats.clear()
        self.total_detections = 0

        # 生成文件名
        timestamp = int(self.start_time)
        self.filename_base = f"{self.camera_name}_{timestamp}"
        self.mp4_path = self.output_dir / f"{self.filename_base}.mp4"
        self.json_path = self.output_dir / f"{self.filename_base}.json"

        # 初始化视频写入器（稍后在添加第一帧时设置）
        self.video_writer = None

    def add_frame(self, frame, detection_info):
        """添加帧"""
        if not self.is_recording:
            return

        # 如果是第一帧，初始化视频写入器
        if self.video_writer is None:
            height, width = frame.shape[:2]
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
            self.video_writer = cv2.VideoWriter(str(self.mp4_path), fourcc, self.fps, (width, height))

        # 写入帧 - 解决色差问题：将RGB转换为BGR
        if frame.shape[2] == 3:  # 确保是3通道彩色图像
            bgr_frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
            self.video_writer.write(bgr_frame)
        else:
            self.video_writer.write(frame)

        self.frames.append(frame.copy())

        # 更新检测统计
        if detection_info and detection_info.get('results'):
            results = detection_info['results']
            if hasattr(results[0], 'boxes') and results[0].boxes and len(results[0].boxes) > 0:
                self.total_detections += len(results[0].boxes)

                # Count classes
                if hasattr(results[0].boxes, 'cls'):
                    classes = results[0].boxes.cls.cpu().numpy().astype(int)
                    class_names = detection_info.get('class_names', [])

                    for cls in classes:
                        if cls < len(class_names):
                            class_name = class_names[cls]
                            self.detection_stats[class_name] = self.detection_stats.get(class_name, 0) + 1

        # Check if need to save file
        if len(self.frames) >= self.max_frames_per_file:
            self.save_recording()
            self.start_recording()  # Start new recording

    def stop_recording(self):
        """Stop recording"""
        if not self.is_recording:
            return

        self.is_recording = False
        self.end_time = time.time()

        if self.video_writer:
            self.video_writer.release()
            self.video_writer = None

        # Save recording
        if self.frames:
            self.save_recording()

    def save_recording(self):
        """Save recording"""
        if not self.frames or not self.start_time:
            return

        # Ensure video writer is released
        if self.video_writer:
            self.video_writer.release()
            self.video_writer = None

        # Save JSON metadata
        metadata = {
            'camera_id': self.camera_id,
            'camera_name': self.camera_name,
            'start_time': self.start_time,
            'end_time': self.end_time or time.time(),
            'fps': self.fps,
            'total_detections': self.total_detections,
            'detection_stats': self.detection_stats,
            'frame_count': len(self.frames),
            'mp4_filename': self.mp4_path.name,
            'json_filename': self.json_path.name
        }

        with open(self.json_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

        print(f"Saving monitoring snapshot: {self.camera_name} - {len(self.frames)} frames, {self.total_detections} detections")
        print(f"File path: {self.mp4_path}")
        print(f"JSON path: {self.json_path}")


class VideoWidget(QWidget):
    """Custom video display component with control functions"""

    def __init__(self, camera_id=0, parent=None):
        super().__init__(parent)
        self.camera_id = camera_id
        self.current_frame = None
        self.detection_state = "NORMAL"  # {0: 'Fall Detected', 1: 'Walking', 2: 'Sitting'}
        self.confidence = 0.0
        self.is_monitoring = False

        self.setup_ui()
        self.setMinimumSize(320, 240)

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)

        # Video display area
        self.video_label = QLabel('Video display area')
        self.video_label.setStyleSheet(StyleManager.get_video_label_style())
        self.video_label.setAlignment(Qt.AlignCenter)
        # Control button area
        control_layout = QHBoxLayout()

        self.start_btn = QPushButton("▶️")
        self.start_btn.setObjectName("startBtn")
        self.start_btn.setToolTip("Start detection")
        self.start_btn.clicked.connect(self.start_detection)

        self.pause_btn = QPushButton("⏸️")
        self.pause_btn.setObjectName("pauseBtn")
        self.pause_btn.setToolTip("Pause detection")
        self.pause_btn.clicked.connect(self.pause_detection)

        self.stop_btn = QPushButton("⏹️")
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.setToolTip("Stop detection")
        self.stop_btn.clicked.connect(self.stop_detection)

        self.monitor_btn = QPushButton("👁️")
        self.monitor_btn.setObjectName("monitorBtn")
        self.monitor_btn.setToolTip("Monitor mode")
        self.monitor_btn.clicked.connect(self.toggle_monitor_mode)

        self.clear_btn = QPushButton("🗑️")
        self.clear_btn.setObjectName("clearBtn")
        self.clear_btn.setToolTip("Clear screen")
        self.clear_btn.clicked.connect(self.clear_frame)

        # 状态标签
        self.status_label = QLabel("🟢 Ready")
        
        control_layout.addWidget(self.start_btn)
        control_layout.addWidget(self.pause_btn)
        control_layout.addWidget(self.stop_btn)
        control_layout.addWidget(self.monitor_btn)
        control_layout.addWidget(self.clear_btn)
        control_layout.addWidget(self.status_label)

        control_layout.addStretch()

        layout.addWidget(self.video_label, stretch=6)
        layout.addLayout(control_layout)
        self.setStyleSheet(StyleManager.get_main_stylesheet())

    def update_frame(self, frame, state="NORMAL", confidence=0.0):
        """更新视频帧"""
        if self.is_monitoring:
            # 监控模式下只显示原始画面
            self.current_frame = frame
            self.display_frame(frame)
            return

        self.current_frame = frame
        self.detection_state = state
        self.confidence = confidence

        # 根据状态设置显示样式  {0: 'Fall Detected', 1: 'Walking', 2: 'Sitting'}
        if state == "Fall Detected":
            self.status_label.setText(f"⚠️ Fall detected (Confidence: {confidence:.2f})")
            self.status_label.setStyleSheet("""
                QLabel {
                    color: #B91C1C;
                    background: rgba(220, 38, 38, 0.1);
                    border: 1px solid #DC2626;
                }
            """)
        elif state == "Walking":
            self.status_label.setText(f"🚶 Walking (Confidence: {confidence:.2f})")
            self.status_label.setStyleSheet("""
                QLabel {
                    color: #2563EB;
                    background: rgba(59, 130, 246, 0.1);
                    border: 1px solid #3B82F6;
                }
            """)
        else:
            self.status_label.setText(f"🪑 Sitting (Confidence: {confidence:.2f})")
            self.status_label.setStyleSheet("""
                QLabel {
                    color: #059669;
                    background: rgba(16, 185, 129, 0.1);
                    border: 1px solid #10B981;
                }
            """)

        self.display_frame(frame)

    def display_frame(self, frame):
        """显示视频帧"""
        if frame is None:
            return

        # 转换OpenCV图像到QImage
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = frame_rgb.shape
        bytes_per_line = ch * w
        q_img = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_img)

        # 缩放图像适应标签大小
        scaled_pixmap = pixmap.scaled(
            self.video_label.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        self.video_label.setPixmap(scaled_pixmap)

    def start_detection(self, camera_id=None):
        """启动检测"""
        if camera_id is not None:
            self.camera_id = camera_id

        self.is_monitoring = False
        self.status_label.setText("🟡 Detecting...")

        # TODO: 实际启动摄像头检测逻辑
        print(f"Start camera detection {self.camera_id}")

    def pause_detection(self):
        """暂停检测"""
        self.status_label.setText("⏸️ Paused")

        # TODO: 实际暂停检测逻辑
        print(f"Pause camera detection {self.camera_id}")

    def stop_detection(self):
        """停止检测"""
        self.status_label.setText("⏹️ Stopped")

        # TODO: 实际停止检测逻辑
        print(f"Stop camera detection {self.camera_id}")

    def toggle_monitor_mode(self):
        """切换监控模式"""
        self.is_monitoring = not self.is_monitoring
        if self.is_monitoring:
            self.monitor_btn.setText("🔍")
            self.monitor_btn.setToolTip("Exit monitor mode")
            self.status_label.setText("👁️ Monitor mode")

        else:
            self.monitor_btn.setText("👁️")
            self.monitor_btn.setToolTip("Monitor mode")
            self.status_label.setText("🟢 Ready")

        print(f"Camera {self.camera_id} monitor mode: {self.is_monitoring}")

    def clear_frame(self):
        """清空画面"""
        self.video_label.clear()
        self.video_label.setText("Camera not active")
        self.status_label.setText("⚪ Free")

        print(f"Clear camera {self.camera_id} screen")

    def set_monitor_mode(self, enable):
        """设置监控模式"""
        self.is_monitoring = enable
        self.toggle_monitor_mode()


import time
import cv2
import numpy as np
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QPushButton,
                               QLabel, QListWidget, QListWidgetItem, QGroupBox, QScrollArea,
                               QMessageBox, QComboBox, QDialog, QFileDialog, QTableWidget,
                               QTableWidgetItem, QHeaderView, QSlider, QSpinBox, QTextEdit)
from PySide6.QtCore import Qt, QTimer, QDateTime, QThread, Signal, Slot, QSize
from PySide6.QtGui import QImage, QPixmap, QFont, QColor
from pathlib import Path
import os
import json


class CameraThread(QThread):
    """摄像头线程，负责捕获和处理视频流"""
    frame_ready = Signal(int, np.ndarray, str, int)  # camera_id, image, status, detection_result

    def __init__(self, camera_id, model=None):
        super().__init__()
        self.camera_id = camera_id
        self.model = model
        self._run_flag = True
        self._paused_flag = False
        self.cap = None

    def run(self):
        """主线程逻辑"""
        try:
            # Explicitly specify CAP_DSHOW backend for Windows stability
            self.cap = cv2.VideoCapture(self.camera_id, cv2.CAP_DSHOW)
            if not self.cap.isOpened():
                self.frame_ready.emit(self.camera_id, None, f"Error: Camera not found {self.camera_id}", -1)
                return

            self.frame_ready.emit(self.camera_id, None, "Status: Running", -1)

            while self._run_flag:
                if not self._paused_flag:
                    ret, frame = self.cap.read()
                    if not ret:
                        self.frame_ready.emit(self.camera_id, None, f"Error: Failed to read camera {self.camera_id}", -1)
                        break

                    # 转换颜色空间 BGR -> RGB
                    rgb_image = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

                    # 如果有模型，进行检测
                    detection_result = -1  # -1表示无检测结果
                    if self.model:
                        # 这里应该是您的模型检测逻辑
                        # 假设模型返回检测结果 (0: 摔倒, 1: 行走, 2: 坐下)
                        # 实际实现需要根据您的模型调整
                        detection_result = self.detect_with_model(rgb_image)

                    self.frame_ready.emit(self.camera_id, rgb_image,
                                          f"Status: Running - {self.get_status_text(detection_result)}",
                                          detection_result)

                # 控制帧率
                time.sleep(0.03)  # ~30fps

        except Exception as e:
            self.frame_ready.emit(self.camera_id, None, f"Error: {str(e)}", -1)
        finally:
            if self.cap:
                self.cap.release()

    def detect_with_model(self, image):
        """使用模型进行检测"""
        # 这里应该是您的实际模型检测代码
        # 返回检测结果 (0: 摔倒, 1: 行走, 2: 坐下)
        # 示例: 随机返回一个结果用于演示
        return np.random.randint(0, 3)

    def get_status_text(self, result):
        """获取状态文本"""
        status_map = {
            -1: "No detection",
            0: "Fall Detected",
            1: "Walking",
            2: "Sitting"
        }
        return status_map.get(result, "Unknown status")

    def stop(self):
        """停止线程"""
        self._run_flag = False
        self.wait()

    def pause(self):
        """暂停线程"""
        self._paused_flag = True

    def resume(self):
        """恢复线程"""
        self._paused_flag = False


class EnhancedMonitoringWidget(QWidget):
    """增强版监控页面组件，支持四分屏动态布局"""

    def __init__(self, model_manager, camera_manager):
        super().__init__()
        self.model_manager = model_manager
        self.camera_manager = camera_manager
        self.camera_threads = {}  # 存储摄像头线程
        self.camera_widgets = {}  # 存储每个摄像头的控件和状态
        self.current_model = None
        self.detection_stats = {0: 0, 1: 0, 2: 0}  # 摔倒检测统计
        self.init_ui()
        self.init_timer()

    def init_timer(self):
        """初始化定时器用于更新时间"""
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_time)
        self.timer.start(1000)  # 每秒更新一次

    def init_ui(self):
        layout = QVBoxLayout(self)

        # 顶部控制区域
        self.init_control_panel(layout)

        # 监控显示区域
        self.init_monitor_area(layout)

    def init_control_panel(self, parent_layout):
        """初始化控制面板"""
        control_group = QGroupBox("🖥️ Monitoring Control")
        control_layout = QHBoxLayout(control_group)

        # 左侧区域：模型和摄像头选择
        left_panel = QVBoxLayout()

        # 模型选择
        model_layout = QHBoxLayout()
        model_layout.addWidget(QLabel("Model:"))

        self.model_combo = QComboBox()
        self.model_combo.currentTextChanged.connect(self.on_model_changed)
        self.init_model_combo()
        model_layout.addWidget(self.model_combo)

        select_model_btn = QPushButton("🔧 Select Model")
        select_model_btn.clicked.connect(self.select_model)
        model_layout.addWidget(select_model_btn)
        left_panel.addLayout(model_layout)

        # 摄像头选择
        camera_layout = QHBoxLayout()
        camera_layout.addWidget(QLabel("Camera:"), stretch=4)

        self.camera_list = QListWidget()
        self.camera_list.setMaximumWidth(300)
        self.camera_list.setSelectionMode(QListWidget.MultiSelection)
        self.refresh_cameras()
        camera_layout.addWidget(self.camera_list)

        refresh_camera_btn = QPushButton("🔄 Refresh")
        refresh_camera_btn.clicked.connect(self.refresh_cameras)
        camera_layout.addWidget(refresh_camera_btn)
        left_panel.addLayout(camera_layout)

        control_layout.addLayout(left_panel)

        # 中间区域：全局控制按钮
        center_panel = QVBoxLayout()
        btn_layout = QHBoxLayout()

        self.start_all_btn = QPushButton("🚀 Start All")
        self.start_all_btn.clicked.connect(self.start_all_cameras)
        btn_layout.addWidget(self.start_all_btn)

        self.pause_all_btn = QPushButton("⏸️ Pause All")
        self.pause_all_btn.clicked.connect(self.pause_all_cameras)
        self.pause_all_btn.setEnabled(False)
        btn_layout.addWidget(self.pause_all_btn)

        self.stop_all_btn = QPushButton("🛑 Stop All")
        self.stop_all_btn.clicked.connect(self.stop_all_cameras)
        self.stop_all_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_all_btn)

        self.clear_all_btn = QPushButton("🗑️ Clear All")
        self.clear_all_btn.clicked.connect(self.clear_all_cameras)
        self.clear_all_btn.setEnabled(False)
        btn_layout.addWidget(self.clear_all_btn)

        center_panel.addLayout(btn_layout)
        control_layout.addLayout(center_panel, stretch=4)

        # Right area: Status information
        right_panel = QVBoxLayout()

        # System time display
        self.time_label = QLabel()
        self.time_label.setAlignment(Qt.AlignRight)
        self.update_time()
        right_panel.addWidget(self.time_label)

        # Detection statistics
        stats_layout = QHBoxLayout()
        stats_layout.addWidget(QLabel("Detection Statistics:"))

        self.fall_label = QLabel("Fall: 0")
        self.walk_label = QLabel("Walk: 0")
        self.sit_label = QLabel("Sit: 0")

        stats_layout.addWidget(self.fall_label)
        stats_layout.addWidget(self.walk_label)
        stats_layout.addWidget(self.sit_label)
        right_panel.addLayout(stats_layout)

        control_layout.addLayout(right_panel)

        parent_layout.addWidget(control_group)

    def init_monitor_area(self, parent_layout):
        """Initialize monitoring display area"""
        self.monitor_scroll = QScrollArea()
        self.monitor_scroll.setStyleSheet("""
            QScrollArea {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 rgba(236, 240, 241, 0.9),
                    stop:1 rgba(189, 195, 199, 0.9));
                border-radius: 8px;
            }
            QScrollArea > QWidget > QWidget {
                background: transparent;
            }
            QScrollArea::corner {
                background: transparent;
            }
        """)

        self.monitor_widget = QWidget()
        self.monitor_layout = QGridLayout(self.monitor_widget)
        self.monitor_scroll.setWidget(self.monitor_widget)
        self.monitor_scroll.setWidgetResizable(True)

        parent_layout.addWidget(self.monitor_scroll)

    def init_model_combo(self):
        """Initialize model dropdown"""
        self.model_combo.clear()
        models = self.model_manager.scan_models()

        if not models:
            self.model_combo.addItem("No models available")
            self.model_combo.setEnabled(False)
        else:
            self.model_combo.addItems([model['name'] for model in models])
            self.model_combo.setEnabled(True)
            self.try_load_default_model()

    def try_load_default_model(self):
        """尝试加载默认模型"""
        if self.model_combo.count() > 0 and self.model_combo.itemText(0) != "No models available":
            first_model = self.model_combo.itemText(0)
            self.load_model_by_name(first_model)

    def load_model_by_name(self, model_name):
        """根据名称加载模型"""
        models = self.model_manager.scan_models()
        for model in models:
            if model['name'] == model_name:
                self.load_model(model['path'])
                break

    def on_model_changed(self, model_text):
        """Model selection changed"""
        if model_text != "No models available":
            self.load_model_by_name(model_text)

    def load_model(self, model_path):
        """加载模型"""
        try:
            self.current_model = YOLO(model_path)
            self.start_all_btn.setEnabled(True)
            return True
        except Exception as e:
            print(f"Model loading failed: {e}")
            return False

    def select_model(self):
        """Select model"""
        dialog = ModelSelectionDialog(self.model_manager, self)
        if dialog.exec() == QDialog.Accepted and dialog.selected_model:
            try:
                self.current_model = YOLO(dialog.selected_model)
                model_name = Path(dialog.selected_model).name
                self.model_combo.clear()
                self.model_combo.addItem(model_name)
                self.start_all_btn.setEnabled(True)
                QMessageBox.information(self, "Success", f"Model loaded successfully: {model_name}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Model loading failed: {str(e)}")

    def refresh_cameras(self):
        """Refresh camera list"""
        self.camera_manager.scan_cameras()
        self.camera_list.clear()

        for camera in self.camera_manager.get_available_cameras():
            item = QListWidgetItem(f"📹 {camera['name']} ({camera['resolution']})")
            item.setData(Qt.UserRole, camera['id'])
            self.camera_list.addItem(item)

    def start_all_cameras(self):
        """Start all selected cameras"""
        if not self.current_model:
            QMessageBox.warning(self, "Warning", "Please select a model first")
            return

        selected_items = self.camera_list.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Warning", "Please select at least one camera")
            return

        camera_ids = [item.data(Qt.UserRole) for item in selected_items]
        self.clear_monitor_display()

        # 根据摄像头数量设置布局
        self.setup_grid_layout(len(camera_ids))

        for cam_id in camera_ids:
            self.add_camera_widget(cam_id)
            self.start_camera_thread(cam_id)

        self.start_all_btn.setEnabled(False)
        self.pause_all_btn.setEnabled(True)
        self.stop_all_btn.setEnabled(True)
        self.clear_all_btn.setEnabled(True)

    def setup_grid_layout(self, num_cameras):
        """根据摄像头数量设置网格布局"""
        # 清除现有布局
        for i in reversed(range(self.monitor_layout.count())):
            widget = self.monitor_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        # 设置新的网格布局
        if num_cameras == 1:
            rows, cols = 1, 1
        elif num_cameras == 2:
            rows, cols = 1, 2
        elif num_cameras == 3:
            rows, cols = 2, 2  # 3个摄像头时使用2x2网格，最后一个位置留空
        else:  # 4个或更多
            rows, cols = 2, 2

        # 设置行列伸缩
        for r in range(rows):
            self.monitor_layout.setRowStretch(r, 1)
        for c in range(cols):
            self.monitor_layout.setColumnStretch(c, 1)

    def add_camera_widget(self, camera_id):
        """为摄像头添加显示和控制部件"""
        camera_group = QGroupBox(f"📹 Camera {camera_id}")
        camera_group.setStyleSheet("""
            QGroupBox {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 rgba(248, 249, 250, 0.9), stop:1 rgba(233, 236, 239, 0.9));
                color: #7f8c8d;
                font-weight: bold;
                font-size: 14px;
                border-radius: 10px;
            }
        """)

        layout = QVBoxLayout(camera_group)

        # 图像显示区域
        image_label = QLabel("Connecting...")
        image_label.setAlignment(Qt.AlignCenter)
        image_label.setStyleSheet("background-color: #ecf0f1;")
        layout.addWidget(image_label)

        # 状态标签
        status_label = QLabel("Status: Connecting...")
        status_label.setStyleSheet("color: #7f8c8d; font-size: 10px;")
        layout.addWidget(status_label)

        # 控制按钮
        btn_layout = QHBoxLayout()

        start_btn = QPushButton("▶️ Start")
        start_btn.clicked.connect(lambda: self.start_camera(camera_id))
        btn_layout.addWidget(start_btn)

        pause_btn = QPushButton("⏸️ 暂停")
        pause_btn.clicked.connect(lambda: self.pause_camera(camera_id))
        pause_btn.setEnabled(False)
        btn_layout.addWidget(pause_btn)

        stop_btn = QPushButton("🛑 Stop")
        stop_btn.clicked.connect(lambda: self.stop_camera(camera_id))
        stop_btn.setEnabled(False)
        btn_layout.addWidget(stop_btn)

        layout.addLayout(btn_layout)

        # 添加到布局
        position = self.get_next_grid_position()
        self.monitor_layout.addWidget(camera_group, position[0], position[1])

        # 保存控件引用
        self.camera_widgets[camera_id] = {
            'group': camera_group,
            'image': image_label,
            'status': status_label,
            'start_btn': start_btn,
            'pause_btn': pause_btn,
            'stop_btn': stop_btn,
            'running': False,
            'paused': False
        }

    def get_next_grid_position(self):
        """获取下一个网格位置"""
        count = len(self.camera_widgets)
        if count == 0:
            return (0, 0)
        elif count == 1:
            return (0, 1)
        elif count == 2:
            return (1, 0)
        elif count == 3:
            return (1, 1)
        else:
            # 超过4个时循环使用位置
            row = (count // 2) % 2
            col = count % 2
            return (row, col)

    def start_camera_thread(self, camera_id):
        """启动摄像头线程"""
        if camera_id in self.camera_widgets and camera_id not in self.camera_threads:
            thread = CameraThread(camera_id, self.current_model)
            thread.frame_ready.connect(self.update_camera_display)
            thread.finished.connect(lambda: self.on_camera_thread_finished(camera_id))
            self.camera_threads[camera_id] = thread
            thread.start()
            self.start_camera(camera_id)

    def start_camera(self, camera_id):
        """启动单个摄像头"""
        if camera_id in self.camera_widgets:
            self.camera_widgets[camera_id]['status'].setText("Status: Running")
            self.camera_widgets[camera_id]['start_btn'].setEnabled(False)
            self.camera_widgets[camera_id]['pause_btn'].setEnabled(True)
            self.camera_widgets[camera_id]['stop_btn'].setEnabled(True)
            self.camera_widgets[camera_id]['running'] = True
            self.camera_widgets[camera_id]['paused'] = False

    def pause_camera(self, camera_id):
        """暂停/继续单个摄像头"""
        if camera_id in self.camera_widgets and camera_id in self.camera_threads:
            widget = self.camera_widgets[camera_id]
            thread = self.camera_threads[camera_id]

            if widget['paused']:
                # 恢复
                thread.resume()
                widget['status'].setText("Status: Running")
                widget['pause_btn'].setText("⏸️ Pause")   
                widget['paused'] = False
            else:
                # 暂停
                thread.pause()
                widget['status'].setText("Status: Paused")
                widget['pause_btn'].setText("▶️ Continue")
                widget['paused'] = True

    def stop_camera(self, camera_id):
        """停止单个摄像头"""
        if camera_id in self.camera_threads:
            self.camera_threads[camera_id].stop()
            self.camera_threads[camera_id].wait()
            del self.camera_threads[camera_id]

        if camera_id in self.camera_widgets:
            self.camera_widgets[camera_id]['status'].setText("Status: Stopped")
            self.camera_widgets[camera_id]['start_btn'].setEnabled(True)
            self.camera_widgets[camera_id]['pause_btn'].setEnabled(False)
            self.camera_widgets[camera_id]['stop_btn'].setEnabled(False)
            self.camera_widgets[camera_id]['running'] = False
            self.camera_widgets[camera_id]['paused'] = False

    def on_camera_thread_finished(self, camera_id):
        """摄像头线程结束时的处理"""
        if camera_id in self.camera_threads:
            del self.camera_threads[camera_id]

        if camera_id in self.camera_widgets:
            self.camera_widgets[camera_id]['status'].setText("Status: Stopped")
            self.camera_widgets[camera_id]['start_btn'].setEnabled(True)
            self.camera_widgets[camera_id]['pause_btn'].setEnabled(False)
            self.camera_widgets[camera_id]['stop_btn'].setEnabled(False)
            self.camera_widgets[camera_id]['running'] = False
            self.camera_widgets[camera_id]['paused'] = False

    def pause_all_cameras(self):
        """暂停/继续所有摄像头"""
        if any(w['running'] for w in self.camera_widgets.values()):
            all_paused = all(w['paused'] for w in self.camera_widgets.values() if w['running'])

            for cam_id, widget in self.camera_widgets.items():
                if widget['running']:
                    if all_paused:
                        # 全部恢复
                        self.pause_camera(cam_id)
                        self.pause_all_btn.setText("⏸️ Pause All")
                    else:
                        # 全部暂停
                        if not widget['paused']:
                            self.pause_camera(cam_id)
                        self.pause_all_btn.setText("▶️ Continue All")

    def stop_all_cameras(self):
        """停止所有摄像头"""
        for cam_id in list(self.camera_threads.keys()):
            self.stop_camera(cam_id)

        self.start_all_btn.setEnabled(True)
        self.pause_all_btn.setEnabled(False)
        self.stop_all_btn.setEnabled(False)

    def clear_all_cameras(self):
        """清除所有摄像头"""
        self.stop_all_cameras()
        self.clear_monitor_display()
        self.start_all_btn.setEnabled(True)
        self.clear_all_btn.setEnabled(False)
        self.detection_stats = {0: 0, 1: 0, 2: 0}
        self.update_stats()

    def clear_monitor_display(self):
        """清空监控显示"""
        for cam_id in list(self.camera_widgets.keys()):
            self.camera_widgets[cam_id]['group'].deleteLater()
        self.camera_widgets.clear()

    def update_time(self):
        """更新时间显示"""
        current_time = QDateTime.currentDateTime().toString("yyyy-MM-dd hh:mm:ss")
        self.time_label.setText(f"🕒 System Time: {current_time}")

    def update_stats(self):
        """更新检测统计"""
        self.fall_label.setText(f"Fall: {self.detection_stats[0]}")
        self.walk_label.setText(f"Walk: {self.detection_stats[1]}")
        self.sit_label.setText(f"Sit: {self.detection_stats[2]}")

    @Slot(int, np.ndarray, str, int)
    def update_camera_display(self, camera_id, image, status, detection_result):
        """更新摄像头显示"""
        if camera_id in self.camera_widgets:
            # 显示图像
            if image is not None:
                height, width, channel = image.shape
                bytes_per_line = 3 * width
                q_image = QImage(image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                pixmap = QPixmap.fromImage(q_image)
                self.camera_widgets[camera_id]['image'].setPixmap(
                    pixmap.scaled(self.camera_widgets[camera_id]['image'].size(),
                                  Qt.KeepAspectRatio, Qt.SmoothTransformation)
                )

            # 更新状态
            if status:
                self.camera_widgets[camera_id]['status'].setText(status)

            # 更新检测统计
            if detection_result in self.detection_stats:
                self.detection_stats[detection_result] += 1
                self.update_stats()


class SliceDetailDialog(QDialog):
    """切片详细信息弹窗"""

    def __init__(self, nii_data, slice_index, direction, parent=None):
        super().__init__(parent)
        self.nii_data = nii_data
        self.current_slice_index = slice_index
        self.direction = direction
        self.max_slices = nii_data.shape[direction]

        self.init_ui()
        self.update_slice_display()
        # 启用鼠标跟踪以捕获滚轮事件
        self.setMouseTracking(True)

    def init_ui(self):
        self.setWindowTitle(f"Slice Detail - Slice {self.current_slice_index}")
        self.resize(600, 600)

        layout = QVBoxLayout(self)

        # 图像显示区域
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setMinimumSize(400, 400)
        # 启用图像标签的鼠标事件
        self.image_label.setMouseTracking(True)
        self.image_label.installEventFilter(self)  # 安装事件过滤器
        layout.addWidget(self.image_label)

        # Control button area
        button_layout = QHBoxLayout()

        self.prev_button = QPushButton("⬆️ Previous")
        self.prev_button.clicked.connect(self.show_previous_slice)
        button_layout.addWidget(self.prev_button)

        self.slice_info_label = QLabel(f"Slice {self.current_slice_index}/{self.max_slices - 1}")
        self.slice_info_label.setAlignment(Qt.AlignCenter)
        button_layout.addWidget(self.slice_info_label)

        self.next_button = QPushButton("⬇️ Next")
        self.next_button.clicked.connect(self.show_next_slice)
        button_layout.addWidget(self.next_button)

        layout.addLayout(button_layout)

        # Close button
        close_button = QPushButton("Close")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)

        # Update button states
        self.update_button_states()

    def eventFilter(self, obj, event):
        """Event filter for handling mouse wheel events"""
        if obj == self.image_label and event.type() == QEvent.Wheel:
            if event.angleDelta().y() > 0:  # Scroll up
                self.show_previous_slice()
            else:  # Scroll down
                self.show_next_slice()
            return True
        return super().eventFilter(obj, event)

    def update_slice_display(self):
        """更新切片显示"""
        try:
            # 提取切片数据
            if self.direction == 0:  # Sagittal
                slice_data = self.nii_data[self.current_slice_index, :, :]
            elif self.direction == 1:  # Coronal
                slice_data = self.nii_data[:, self.current_slice_index, :]
            else:  # Axial
                slice_data = self.nii_data[:, :, self.current_slice_index]

            # 确保数据是连续的
            if not slice_data.flags['C_CONTIGUOUS']:
                slice_data = np.ascontiguousarray(slice_data)

            # 转换为 QImage 显示
            # 归一化数据到 0-255 范围
            slice_normalized = ((slice_data - slice_data.min()) /
                                (slice_data.max() - slice_data.min()) * 255).astype(np.uint8)

            height, width = slice_normalized.shape
            bytes_per_line = width
            q_img = QImage(slice_normalized.data, width, height, bytes_per_line, QImage.Format_Grayscale8)

            # 缩放图像以适应显示区域
            pixmap = QPixmap.fromImage(q_img)
            scaled_pixmap = pixmap.scaled(
                self.image_label.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.image_label.setPixmap(scaled_pixmap)

            # 更新切片信息
            self.slice_info_label.setText(f"Slice {self.current_slice_index}/{self.max_slices - 1}")

            # 更新按钮状态
            self.update_button_states()

        except Exception as e:
            self.image_label.setText(f"Display Error: {str(e)}")

    def update_button_states(self):
        """更新按钮状态"""
        self.prev_button.setEnabled(bool(self.current_slice_index > 0))
        self.next_button.setEnabled(bool(self.current_slice_index < self.max_slices - 1))

    def show_previous_slice(self):
        """显示上一张切片"""
        if self.current_slice_index > 0:
            self.current_slice_index -= 1
            self.update_slice_display()

    def show_next_slice(self):
        """显示下一张切片"""
        if self.current_slice_index < self.max_slices - 1:
            self.current_slice_index += 1
            self.update_slice_display()


class SnapshotWidget(QWidget):
    """监控快照组件 - 用于显示和回放已保存的监控快照"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.snapshots = []  # 存储快照记录
        self.current_snapshot_index = 0
        self.monitor_history_dir = get_runtime_dir("monitor_history")
        self.monitor_history_dir.mkdir(exist_ok=True)
        # 添加detection_history目录支持
        self.detection_history_dir = get_runtime_dir("detection_history")
        self.detection_history_dir.mkdir(exist_ok=True)

        self.init_ui()
        self.load_snapshots()

    def init_ui(self):
        """初始化UI界面"""
        layout = QVBoxLayout(self)

        # Snapshot list and playback area
        content_layout = QHBoxLayout()

        # Left panel: Snapshot list
        left_panel = QVBoxLayout()

        list_group = QGroupBox("📋 Snapshot History")
        list_group.setMaximumHeight(780)
        list_layout = QVBoxLayout(list_group)

        self.snapshot_list = QListWidget()
        self.snapshot_list.itemClicked.connect(self.on_snapshot_selected)
        list_layout.addWidget(self.snapshot_list)

        # Snapshot operation buttons
        snapshot_btn_layout = QHBoxLayout()

        self.play_btn = QPushButton("▶️ Play")
        self.play_btn.clicked.connect(self.play_selected_snapshot)
        self.play_btn.setEnabled(False)
        snapshot_btn_layout.addWidget(self.play_btn)

        self.delete_btn = QPushButton("🗑️ Delete")
        self.delete_btn.clicked.connect(self.delete_selected_snapshot)
        self.delete_btn.setEnabled(False)
        snapshot_btn_layout.addWidget(self.delete_btn)

        self.refresh_btn = QPushButton("🔄 Refresh")
        self.refresh_btn.clicked.connect(self.load_snapshots)
        snapshot_btn_layout.addWidget(self.refresh_btn)

        # Add export button to layout
        self.export_btn = QPushButton("📤 Export")
        self.export_btn.clicked.connect(self.export_selected_snapshot)
        self.export_btn.setEnabled(False)
        snapshot_btn_layout.addWidget(self.export_btn)

        list_layout.addLayout(snapshot_btn_layout)
        left_panel.addWidget(list_group)

        content_layout.addLayout(left_panel, 1)

        # Right panel: Playback area
        right_panel = QVBoxLayout()

        player_group = QGroupBox("🎥 Snapshot Player")
        # player_group.setMaximumHeight(780)  # Reduce height
        player_layout = QVBoxLayout(player_group)

        # Video display area
        self.video_label = QLabel("Select snapshot to play")
        self.video_label.setMinimumSize(640, 390)
        self.video_label.setStyleSheet("""
            QLabel {
                border: 1px solid rgba(52, 152, 219, 0.3);
                font-size: 14px;
                border-radius: 10px;
            }
        """)
        self.video_label.setAlignment(Qt.AlignCenter)
        self.video_label.setScaledContents(True)
        player_layout.addWidget(self.video_label)

        # 播放控制
        playback_layout = QHBoxLayout()

        self.playback_btn = QPushButton("▶️")
        self.playback_btn.clicked.connect(self.toggle_playback)
        self.playback_btn.setEnabled(False)
        playback_layout.addWidget(self.playback_btn)

        self.stop_btn = QPushButton("⏹️")
        self.stop_btn.clicked.connect(self.stop_playback)
        self.stop_btn.setEnabled(False)
        playback_layout.addWidget(self.stop_btn)

        # 进度条
        self.progress_slider = QSlider(Qt.Horizontal)
        self.progress_slider.setEnabled(False)
        self.progress_slider.valueChanged.connect(self.on_progress_changed)
        playback_layout.addWidget(self.progress_slider)

        # 时间显示
        self.time_label = QLabel("00:00 / 00:00")
        self.time_label.setStyleSheet("color: #7f8c8d; font-size: 12px;")
        playback_layout.addWidget(self.time_label)

        player_layout.addLayout(playback_layout, stretch=5)

        # Snapshot information
        info_group = QGroupBox("📊 Snapshot Information")
        info_group.setStyleSheet("""
                    QGroupBox {
                border: 1px solid rgba(52, 152, 219, 0.3);
                font-size: 14px;
                border-radius: 10px;
            }
        """
                                 )
        info_layout = QVBoxLayout(info_group)

        self.info_text = QTextEdit()
        self.info_text.setMinimumHeight(170)
        self.info_text.setReadOnly(True)
        self.info_text.setStyleSheet("""
            QTextEdit {
                background: rgba(248, 249, 250, 0.8);
                border: 1px solid rgba(189, 195, 199, 0.3);
                border-radius: 5px;
                font-size: 11px;
                color: #2c3e50;
            }
        """)
        info_layout.addWidget(self.info_text)

        player_layout.addWidget(info_group, stretch=3)
        right_panel.addWidget(player_group)

        content_layout.addLayout(right_panel, 2)
        layout.addLayout(content_layout)

        # 初始化播放定时器
        self.playback_timer = QTimer()
        self.playback_timer.timeout.connect(self.update_playback)
        self.current_frame_index = 0
        self.is_playing = False

    def toggle_recording(self):
        """Toggle recording status"""
        if not self.is_recording:
            self.start_recording()
        else:
            self.stop_recording()

    def start_recording(self):
        """Start recording"""
        self.is_recording = True
        self.recording_frames.clear()
        self.recording_start_time = time.time()
        self.max_recording_duration = self.duration_spinbox.value()

        self.record_btn.setText("⏹️ Stop Recording")
        self.record_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #95a5a6, stop:1 #7f8c8d);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #7f8c8d, stop:1 #6c7b7d);
            }
        """)
        self.save_btn.setEnabled(False)
        self.clear_btn.setEnabled(True)
        self.recording_status.setText("Status: Recording...")
        self.recording_status.setStyleSheet("""
            QLabel {
                color: #e74c3c;
                font-size: 12px;
                padding: 5px;
                background: rgba(231, 76, 60, 0.1);
                border-radius: 5px;
            }
        """)

    def stop_recording(self):
        """Stop recording"""
        self.is_recording = False

        self.record_btn.setText("🔴 Start Recording")
        self.record_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #e74c3c, stop:1 #c0392b);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #c0392b, stop:1 #a93226);
            }
        """)

        if len(self.recording_frames) > 0:
            self.save_btn.setEnabled(True)

        self.recording_status.setText(f"Status: Recording completed ({len(self.recording_frames)} frames)")
        self.recording_status.setStyleSheet("""
            QLabel {
                color: #27ae60;
                font-size: 12px;
                padding: 5px;
                background: rgba(39, 174, 96, 0.1);
                border-radius: 5px;
            }
        """)

    def add_frame(self, frame, detection_info=None):
        """Add frame to recording"""
        if not self.is_recording:
            return

        # Check recording duration
        if time.time() - self.recording_start_time > self.max_recording_duration:
            self.stop_recording()
            return

        frame_data = {
            'frame': frame.copy(),
            'timestamp': time.time(),
            'detection_info': detection_info or {}
        }
        self.recording_frames.append(frame_data)

        # Update status
        elapsed = time.time() - self.recording_start_time
        self.recording_status.setText(f"Status: Recording... ({len(self.recording_frames)} frames, {elapsed:.1f}s)")

    def save_current_recording(self):
        """Save current recording"""
        if not self.recording_frames:
            QMessageBox.warning(self, "Warning", "No recording content to save")
            return

        # Generate snapshot ID
        snapshot_id = f"snapshot_{int(time.time())}"
        snapshot_path = self.snapshots_dir / f"{snapshot_id}.json"

        # Save frame data
        snapshot_data = {
            'id': snapshot_id,
            'created_time': time.time(),
            'duration': self.max_recording_duration,
            'fps': self.fps_spinbox.value(),
            'frame_count': len(self.recording_frames),
            'frames': []
        }

        # Compress and save frame data
        for i, frame_data in enumerate(self.recording_frames):
            # Convert frame to hex string
            _, buffer = cv2.imencode('.jpg', frame_data['frame'], [cv2.IMWRITE_JPEG_QUALITY, 80])
            frame_str = buffer.tobytes().hex()

            snapshot_data['frames'].append({
                'index': i,
                'timestamp': frame_data['timestamp'],
                'frame_data': frame_str,
                'detection_info': frame_data['detection_info']
            })

        # Save to file
        with open(snapshot_path, 'w', encoding='utf-8') as f:
            json.dump(snapshot_data, f, ensure_ascii=False, indent=2)

        # Add to snapshot list
        snapshot_info = {
            'id': snapshot_id,
            'path': str(snapshot_path),
            'created_time': snapshot_data['created_time'],
            'duration': snapshot_data['duration'],
            'frame_count': snapshot_data['frame_count'],
            'fps': snapshot_data['fps']
        }

        self.snapshots.append(snapshot_info)
        self.update_snapshot_list()

        QMessageBox.information(self, "Success", f"Snapshot saved: {snapshot_id}")

        # 清空当前录制
        self.clear_recording()

    def clear_recording(self):
        """清空当前录制"""
        self.recording_frames.clear()
        self.save_btn.setEnabled(False)
        self.clear_btn.setEnabled(False)
        self.recording_status.setText("Status: Not Recording")
        self.recording_status.setStyleSheet("""
            QLabel {
                color: #7f8c8d;
                font-size: 12px;
                padding: 5px;
                background: rgba(236, 240, 241, 0.5);
                border-radius: 5px;
            }
        """)

    def load_snapshots(self):
        """加载已保存的快照"""
        self.snapshots.clear()

        # 扫描monitor_history目录下的JSON文件
        for json_file in self.monitor_history_dir.glob("*.json"):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # 检查对应的MP4文件是否存在
                mp4_file = json_file.with_suffix('.mp4')
                if mp4_file.exists():
                    snapshot_info = {
                        'camera_name': data.get('camera_name', 'Unknown Camera'),
                        'start_time': data.get('start_time', 0),
                        'end_time': data.get('end_time', 0),
                        'file_size': self._get_file_size(mp4_file),
                        'detection_stats': data.get('detection_stats', {}),
                        'json_path': str(json_file),
                        'mp4_path': str(mp4_file),
                        'fps': data.get('fps', 20),
                        'total_detections': data.get('total_detections', 0),
                        'source': 'monitor'  # 标记来源为监控
                    }
                    self.snapshots.append(snapshot_info)
            except Exception as e:
                print(f"Loading snapshot failed {json_file}: {e}")

        # 扫描detection_history目录下的JSON文件
        for json_file in self.detection_history_dir.glob("*.json"):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # 检查对应的MP4文件是否存在
                mp4_file = json_file.with_suffix('.mp4')
                if mp4_file.exists():
                    snapshot_info = {
                        'camera_name': data.get('source_name', 'Unknown Source'),
                        'start_time': data.get('start_time', 0),
                        'end_time': data.get('end_time', 0),
                        'file_size': self._get_file_size(mp4_file),
                        'detection_stats': data.get('detection_stats', {}),
                        'json_path': str(json_file),
                        'mp4_path': str(mp4_file),
                        'fps': data.get('fps', 20),
                        'total_detections': data.get('total_detections', 0),
                        'source': 'detection'  # 标记来源为检测
                    }
                    self.snapshots.append(snapshot_info)
            except Exception as e:
                print(f"Loading snapshot failed {json_file}: {e}")

        self.update_snapshot_list()

    def _get_file_size(self, file_path):
        """获取文件大小"""
        try:
            size = file_path.stat().st_size
            for unit in ['B', 'KB', 'MB', 'GB']:
                if size < 1024.0:
                    return f"{size:.1f} {unit}"
                size /= 1024.0
            return f"{size:.1f} TB"
        except:
            return "Unknown"

    def update_snapshot_list(self):
        """更新快照列表显示"""
        self.snapshot_list.clear()

        # 按开始时间排序（最新的在前）
        self.snapshots.sort(key=lambda x: x['start_time'], reverse=True)

        for snapshot in self.snapshots:
            start_time = datetime.fromtimestamp(snapshot['start_time'])
            end_time = datetime.fromtimestamp(snapshot['end_time'])

            # 格式化检测统计信息
            stats_text = ""
            if snapshot['detection_stats']:
                stats_items = []
                for class_name, count in snapshot['detection_stats'].items():
                    stats_items.append(f"{class_name}:{count}")
                stats_text = " | ".join(stats_items)

            # 根据来源添加不同的前缀标识
            source_prefix = "🖥️" if snapshot['source'] == 'monitor' else "📹"
            item_text = f"{source_prefix} {snapshot['camera_name']}\n"
            item_text += f"Time: {start_time.strftime('%Y-%m-%d %H:%M:%S')} - {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            item_text += f": {snapshot['file_size']} | Detection: {snapshot['total_detections']}\n"
            if stats_text:
                item_text += f"Detection: {stats_text}"

            item = QListWidgetItem(item_text)
            item.setData(Qt.UserRole, snapshot)
            self.snapshot_list.addItem(item)

    def on_snapshot_selected(self, item):
        """快照被选中"""
        snapshot = item.data(Qt.UserRole)
        self.current_snapshot_index = self.snapshots.index(snapshot)
        self.play_btn.setEnabled(True)
        # self.playback_btn.setEnabled(True)
        self.delete_btn.setEnabled(True)
        self.export_btn.setEnabled(True)

        # 显示快照信息
        self.show_snapshot_info(snapshot)

    def show_snapshot_info(self, snapshot):
        """显示快照信息"""
        start_time = datetime.fromtimestamp(snapshot['start_time'])
        end_time = datetime.fromtimestamp(snapshot['end_time'])
        duration = snapshot['end_time'] - snapshot['start_time']

        info_text = f"Camera: {snapshot['camera_name']}\n"
        info_text += f"StartTime: {start_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        info_text += f"EndTime: {end_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
        info_text += f"Duration: {duration:.1f} s\n"
        info_text += f"Size: {snapshot['file_size']}\n"
        info_text += f"FPS: {snapshot['fps']} fps\n"
        info_text += f"Detection: {snapshot['total_detections']}\n"

        if snapshot['detection_stats']:
            info_text += f"Detection:\n"
            for class_name, count in snapshot['detection_stats'].items():
                info_text += f"  {class_name}: {count} times\n"

        info_text += f"Video File: {snapshot['mp4_path']}"

        self.info_text.setText(info_text)

    def play_selected_snapshot(self):
        """播放选中的快照"""
        if not self.snapshots or self.current_snapshot_index >= len(self.snapshots):
            return

        snapshot = self.snapshots[self.current_snapshot_index]

        try:
            # Use CAP_FFMPEG for video files (more compatible)
            cap = cv2.VideoCapture(snapshot['mp4_path'], cv2.CAP_FFMPEG)
            if not cap.isOpened():
                QMessageBox.warning(self, "Error", "Failed to open video file")
                return

            # 读取所有帧
            self.playback_frames = []
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                self.playback_frames.append(frame)

            cap.release()

            if not self.playback_frames:
                QMessageBox.warning(self, "Error", "Video file is empty")
                return

            # 设置播放参数
            self.current_frame_index = 0
            self.playback_fps = snapshot['fps']
            self.playback_interval = 1000 // self.playback_fps  # 毫秒

            # 设置进度条
            self.progress_slider.setRange(0, len(self.playback_frames) - 1)
            self.progress_slider.setValue(0)
            self.progress_slider.setEnabled(True)
            self.playback_btn.setEnabled(True)

            # 开始播放
            self.toggle_playback()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to play snapshot video: {str(e)}")

    def toggle_playback(self):
        """切换播放状态"""
        if not hasattr(self, 'playback_frames') or not self.playback_frames:
            return

        if self.is_playing:
            self.pause_playback()
        else:
            self.start_playback()

    def start_playback(self):
        """开始播放"""
        self.is_playing = True
        self.playback_btn.setText("⏸️")
        self.stop_btn.setEnabled(True)

        self.playback_timer.start(self.playback_interval)

    def pause_playback(self):
        """暂停播放"""
        self.is_playing = False
        self.playback_btn.setText("▶️")

        self.playback_timer.stop()

    def stop_playback(self):
        """停止播放"""
        self.is_playing = False
        self.playback_btn.setText("▶️")
        self.stop_btn.setEnabled(False)

        self.playback_timer.stop()
        self.current_frame_index = 0
        self.progress_slider.setValue(0)

        # 显示第一帧
        if hasattr(self, 'playback_frames') and self.playback_frames:
            self.display_frame(self.playback_frames[0])

    def update_playback(self):
        """更新播放"""
        if not hasattr(self, 'playback_frames') or not self.playback_frames:
            return

        if self.current_frame_index >= len(self.playback_frames):
            self.stop_playback()
            return

        # 显示当前帧
        frame = self.playback_frames[self.current_frame_index]
        self.display_frame(frame)

        # 更新进度
        self.progress_slider.setValue(self.current_frame_index)

        # 更新时间显示
        current_time = self.current_frame_index / self.playback_fps
        total_time = len(self.playback_frames) / self.playback_fps
        self.time_label.setText(f"{current_time:.1f}s / {total_time:.1f}s")

        self.current_frame_index += 1

    def on_progress_changed(self, value):
        """进度条改变"""
        if hasattr(self, 'playback_frames') and self.playback_frames and not self.is_playing:
            self.current_frame_index = value
            frame = self.playback_frames[value]
            self.display_frame(frame)

            # 更新时间显示
            current_time = value / self.playback_fps
            total_time = len(self.playback_frames) / self.playback_fps
            self.time_label.setText(f"{current_time:.1f}s / {total_time:.1f}s")

    def display_frame(self, frame):
        """显示帧"""
        if frame is None:
            return

        # 转换颜色空间
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        h, w, ch = frame_rgb.shape
        bytes_per_line = ch * w
        q_image = QImage(frame_rgb.data, w, h, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_image)

        # 缩放适应显示区域
        scaled_pixmap = pixmap.scaled(
            self.video_label.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation
        )
        self.video_label.setPixmap(scaled_pixmap)

    def delete_selected_snapshot(self):
        """Delete selected snapshot"""
        if not self.snapshots or self.current_snapshot_index >= len(self.snapshots):
            return

        snapshot = self.snapshots[self.current_snapshot_index]

        reply = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete the snapshot for camera '{snapshot['camera_name']}'?\nThis operation cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # Delete MP4 and JSON files
                Path(snapshot['mp4_path']).unlink()
                Path(snapshot['json_path']).unlink()

                # Remove from list
                self.snapshots.pop(self.current_snapshot_index)
                self.update_snapshot_list()

                # Clear playback area
                self.video_label.clear()
                self.video_label.setText("Select snapshot to play")
                self.info_text.clear()

                # Disable buttons
                self.play_btn.setEnabled(False)
                self.delete_btn.setEnabled(False)
                self.export_btn.setEnabled(False)

                QMessageBox.information(self, "Success", "Snapshot deleted")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to delete snapshot: {str(e)}")

    # 添加导出功能的实现
    def export_selected_snapshot(self):
        """Export selected snapshot"""
        if not self.snapshots or self.current_snapshot_index >= len(self.snapshots):
            return

        snapshot = self.snapshots[self.current_snapshot_index]

        # Select export directory
        export_dir = QFileDialog.getExistingDirectory(self, "Select Export Directory")
        if not export_dir:
            return

        try:
            export_path = Path(export_dir)
            mp4_file = Path(snapshot['mp4_path'])
            json_file = Path(snapshot['json_path'])

            # Construct export file paths
            mp4_export_path = export_path / mp4_file.name
            json_export_path = export_path / json_file.name

            # Copy files
            import shutil
            shutil.copy2(mp4_file, mp4_export_path)
            shutil.copy2(json_file, json_export_path)

            QMessageBox.information(self, "Success", f"Snapshot exported to:\n{export_path}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export snapshot: {str(e)}")


class EnhancedDetectionUI(QMainWindow):
    """Enhanced detection UI main window"""

    def __init__(self):
        super().__init__()
        self.model = None
        self.detection_thread = None
        self.batch_detection_thread = None
        self.current_source_type = 'image'
        self.current_source_path = None
        self.confidence_threshold = 0.25
        self.batch_results = []
        self.current_batch_index = 0

        # Snapshot related properties
        self.is_auto_saving = False
        self.video_recorder = None
        self.history_dir = get_runtime_dir("detection_history")
        self.history_dir.mkdir(exist_ok=True)

        # Managers
        self.camera_manager = CameraManager()
        self.model_manager = ModelManager()
        self.log_text = QTextEdit()
        self.init_ui()
        # Set window icon from logo.ico file (compatible with PyInstaller)
        icon_path = get_resource_path("img/logo.ico")
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))

        # Apply style
        self.setStyleSheet(StyleManager.get_main_stylesheet())
        self.setup_title_shortcut()

        self.slice_update_timer = QTimer()
        self.slice_update_timer.setSingleShot(True)
        self.slice_update_timer.timeout.connect(self.update_slice_preview)
        self.slice_range_changed = False

    def init_ui(self):
        """Initialize UI"""
        self.setWindowTitle("🚀 MediScreen-Brain Tumor Detection System ")
        self.setGeometry(100, 100, 1400, 750)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Create main splitter
        main_splitter = QSplitter(Qt.Horizontal)

        # Left control panel
        left_widget = self.create_control_panel()
        left_widget.setMaximumWidth(500)
        left_widget.setMinimumWidth(400)

        # Right display area
        right_widget = self.create_display_area()

        main_splitter.addWidget(left_widget)
        main_splitter.addWidget(right_widget)
        main_splitter.setSizes([450, 1250])

        main_layout.addWidget(main_splitter)

        # Status bar
        self.statusBar().showMessage("🎯 Ready - Please select model and detection source")

    def setup_title_shortcut(self):
        """Set title edit shortcut"""
        title_shortcut = QShortcut(QKeySequence("F2"), self)
        title_shortcut.activated.connect(self.edit_window_title)
        # Add new Ctrl+R shortcut
        title_shortcut_ctrl_r = QShortcut(QKeySequence("Ctrl+R"), self)
        title_shortcut_ctrl_r.activated.connect(self.edit_window_title)

    def edit_window_title(self):
        """Edit window title"""
        current_title = self.windowTitle().strip()
        new_title, ok = QInputDialog.getText(
            self,
            "Edit Window Title",
            "Please enter new window title:",
            text=current_title
        )

        if ok and new_title:
            self.setWindowTitle(new_title)

    def create_control_panel(self):
        """Create control panel"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Model configuration
        model_group = QGroupBox("🤖 Model Configuration")
        model_layout = QVBoxLayout(model_group)

        # Model selection
        model_select_layout = QHBoxLayout()
        model_select_layout.addWidget(QLabel("Select Model:"))

        self.model_combo = QComboBox()
        self.model_combo.currentTextChanged.connect(self.on_model_changed)
        self.init_model_combo()
        model_select_layout.addWidget(self.model_combo)

        advanced_model_btn = QPushButton("🔧")
        advanced_model_btn.clicked.connect(self.show_model_selection_dialog)
        advanced_model_btn.setMaximumWidth(80)
        model_select_layout.addWidget(advanced_model_btn)

        model_layout.addLayout(model_select_layout)

        # Confidence Configuration
        conf_layout = QHBoxLayout()
        conf_layout.addWidget(QLabel("Confidence Threshold:"))

        self.conf_slider = QSlider(Qt.Horizontal)
        self.conf_slider.setMinimum(1)
        self.conf_slider.setMaximum(100)
        self.conf_slider.setValue(25)
        self.conf_slider.valueChanged.connect(self.on_confidence_changed)
        conf_layout.addWidget(self.conf_slider)

        self.conf_spinbox = QDoubleSpinBox()
        self.conf_spinbox.setRange(0.01, 1.0)
        self.conf_spinbox.setSingleStep(0.01)
        self.conf_spinbox.setValue(0.25)
        self.conf_spinbox.setDecimals(2)
        self.conf_spinbox.valueChanged.connect(self.on_confidence_spinbox_changed)
        conf_layout.addWidget(self.conf_spinbox)

        model_layout.addLayout(conf_layout)
        layout.addWidget(model_group)

        # Detection Source Configuration
        source_group = QGroupBox("📁 Detection Source Configuration")
        source_layout = QVBoxLayout(source_group)

        # Detection Mode Selection
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel("Detection Mode:"))

        self.source_combo = QComboBox()
        self.source_combo.addItems(["📷 Single Image", "🎬 Video File", "📹 Camera", "📂 Batch Folder", "🧠 NIfTI File"])
        self.source_combo.currentTextChanged.connect(self.on_source_changed)
        mode_layout.addWidget(self.source_combo)
        source_layout.addLayout(mode_layout)

        # Camera Selection (only visible in camera mode)
        self.camera_select_layout = QHBoxLayout()
        self.camera_select_layout.addWidget(QLabel("Camera:"))

        self.camera_combo = QComboBox()
        self.refresh_camera_list()
        self.camera_select_layout.addWidget(self.camera_combo)

        refresh_camera_btn = QPushButton("🔄")
        refresh_camera_btn.setMaximumWidth(40)
        refresh_camera_btn.clicked.connect(self.refresh_camera_list)
        self.camera_select_layout.addWidget(refresh_camera_btn)

        source_layout.addLayout(self.camera_select_layout)

        # File Selection
        file_layout = QHBoxLayout()
        self.select_file_btn = QPushButton("📁 Select File/Folder")
        self.select_file_btn.clicked.connect(self.select_file)
        file_layout.addWidget(self.select_file_btn)
        source_layout.addLayout(file_layout)

        # Current File Display
        self.current_file_label = QLabel("No file selected")
        self.current_file_label.setWordWrap(True)
        self.current_file_label.setStyleSheet("color: #7f8c8d; font-size: 11px; padding: 5px;")
        source_layout.addWidget(self.current_file_label)

        layout.addWidget(source_group)

        # Detection Control
        control_group = QGroupBox("🎮 Detection Control")
        control_layout = QVBoxLayout(control_group)

        # Control Buttons
        btn_layout = QHBoxLayout()

        self.start_btn = QPushButton("▶️ Start")
        self.start_btn.clicked.connect(self.start_detection)
        self.start_btn.setEnabled(False)
        btn_layout.addWidget(self.start_btn)

        self.pause_btn = QPushButton("⏸️ Pause")
        self.pause_btn.clicked.connect(self.pause_detection)
        self.pause_btn.setEnabled(False)
        btn_layout.addWidget(self.pause_btn)

        self.stop_btn = QPushButton("⏹️ Stop")
        self.stop_btn.clicked.connect(self.stop_detection)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        self.video_is_auto_saving = False
        self.kuaizhao_btn = QPushButton("🎬 Snapshot")
        self.kuaizhao_btn.clicked.connect(self.kuaizhao_detection)
        self.kuaizhao_btn.setEnabled(False)
        btn_layout.addWidget(self.kuaizhao_btn)

        control_layout.addLayout(btn_layout)

        # Progress Bar
        progress_layout = QHBoxLayout()
        progress_layout.addWidget(QLabel("Progress:"))

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        progress_layout.addWidget(self.progress_bar)

        control_layout.addLayout(progress_layout)

        layout.addWidget(control_group)

        # Log Area
        log_group = QGroupBox("📋 Running Log")
        log_layout = QVBoxLayout(log_group)

        self.log_text.setMinimumHeight(180)
        self.log_text.setFont(QFont("Consolas", 10))
        log_layout.addWidget(self.log_text)

        log_btn_layout = QHBoxLayout()
        log_btn_layout.addStretch()

        self.clear_log_btn = QPushButton("🗑️ Clear")
        self.clear_log_btn.clicked.connect(self.clear_log)
        self.clear_log_btn.setMaximumWidth(100)
        log_btn_layout.addWidget(self.clear_log_btn)

        log_layout.addLayout(log_btn_layout)
        layout.addWidget(log_group)

        # layout.addStretch()
        return widget

    def create_display_area(self):
        """Create display area"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Create tabs
        self.tab_widget = QTabWidget()

        # Real-time detection tab
        realtime_tab = self.create_realtime_tab
        self.tab_widget.addTab(realtime_tab, "🎯 Real-time Detection")
        # Create NIfTI tumor detection tab
        niigz_detection_tab = self.create_niigz_detection_tab()
        self.tab_widget.addTab(niigz_detection_tab, "🧠 NIfTI Detection")
        # Batch results tab
        batch_tab = self.create_batch_tab()
        self.tab_widget.addTab(batch_tab, "📊 Batch Results")

        # Monitoring page tab
        monitor_tab = MonitoringWidget(self.model_manager, self.camera_manager)
        self.tab_widget.addTab(monitor_tab, "🖥️ Real-time Monitoring")

        # Monitoring snapshot tab
        self.snapshot_widget = SnapshotWidget()
        self.tab_widget.addTab(self.snapshot_widget, "🎬 Monitoring Snapshot")

        # Create NIfTI format conversion tab
        nifti_tab = self.create_nifti_conversion_tab()
        self.tab_widget.addTab(nifti_tab, " NIfTI Conversion")



        layout.addWidget(self.tab_widget)
        return widget

    def create_niigz_detection_tab(self):
        """Create NIfTI tumor detection tab"""
        tab = QWidget()
        tab.setMaximumWidth(1200)
        layout = QVBoxLayout(tab)

        # Three-axis display area
        axes_layout = QHBoxLayout()
        axes_layout.setSpacing(2)
        axes_layout.setContentsMargins(0, 5, 0, 8)
        # Axial view
        axial_widget = QWidget()
        axial_vlayout = QVBoxLayout(axial_widget)
        axial_vlayout.setContentsMargins(0, 0, 0, 0)
        axial_label = QLabel("Axial")
        axial_label.setAlignment(Qt.AlignCenter)
        axial_label.setStyleSheet("font-weight: bold;")
        axial_vlayout.addWidget(axial_label)
        self.axial_display = QLabel("Not detected")
        self.axial_display.setAlignment(Qt.AlignCenter)
        self.axial_display.setStyleSheet(StyleManager.get_image_niigz_label_style())
        self.axial_display.setMinimumHeight(300)
        axial_vlayout.addWidget(self.axial_display)
        axes_layout.addWidget(axial_widget)

        # Sagittal view
        sagittal_widget = QWidget()
        sagittal_vlayout = QVBoxLayout(sagittal_widget)
        sagittal_vlayout.setContentsMargins(0, 0, 0, 0)
        sagittal_label = QLabel("Sagittal")
        sagittal_label.setAlignment(Qt.AlignCenter)
        sagittal_label.setStyleSheet("font-weight: bold;")
        sagittal_vlayout.addWidget(sagittal_label)
        self.sagittal_display = QLabel("Not detected")
        self.sagittal_display.setAlignment(Qt.AlignCenter)
        self.sagittal_display.setStyleSheet(StyleManager.get_image_niigz_label_style())
        self.sagittal_display.setMinimumHeight(300)
        sagittal_vlayout.addWidget(self.sagittal_display)
        axes_layout.addWidget(sagittal_widget)

        # Coronal view
        coronal_widget = QWidget()
        coronal_vlayout = QVBoxLayout(coronal_widget)
        coronal_vlayout.setContentsMargins(0, 0, 0, 0)
        coronal_label = QLabel("Coronal")
        coronal_label.setAlignment(Qt.AlignCenter)
        coronal_label.setStyleSheet("font-weight: bold;")
        coronal_vlayout.addWidget(coronal_label)
        self.coronal_display = QLabel("Not detected")
        self.coronal_display.setAlignment(Qt.AlignCenter)
        self.coronal_display.setStyleSheet(StyleManager.get_image_niigz_label_style())
        self.coronal_display.setMinimumHeight(300)
        coronal_vlayout.addWidget(self.coronal_display)
        axes_layout.addWidget(coronal_widget)

        layout.addLayout(axes_layout)

        # Detection Information Table
        table_layout = QVBoxLayout()
        title = QLabel("🎯 Detection Results Detail Table")
        title.setStyleSheet("font-size: 16px; font-weight: bold; color: #2c3e50; margin-bottom: 5px;")
        table_layout.addWidget(title)
        self.niigz_table = QTableWidget()
        self.niigz_table.setColumnCount(6)
        self.niigz_table.setHorizontalHeaderLabels(["Slice", "Class", "Confidence", "Coordinates", "Diameter(mm)", "Area(mm²)"])
        self.niigz_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.niigz_table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: white;  /* Deep blue background */
                color: #2c3e50;               
                font-size: 8pt;
                font-weight: bold;
                height: 12px;
                border: 1px solid #d0d0d0;       /* Border */
            }
        """)
        table_layout.addWidget(self.niigz_table)

        layout.addLayout(table_layout)

        # Detection Summary
        self.niigz_summary_label = QLabel("Detection Summary")
        self.niigz_summary_label.setMinimumHeight(80)
        # self.niigz_summary_label.setAlignment(Qt.AlignCenter)
        self.niigz_summary_label.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(236, 240, 241, 0.9), stop:1 rgba(189, 195, 199, 0.9));
            padding: 12px;
            border-radius: 8px;
            font-size: 12px;
            color: #2c3e50;
            font-weight: bold;
        """)
        layout.addWidget(self.niigz_summary_label)

        bututton_layout = QHBoxLayout()
        bututton_layout.addStretch()
        self.select_target_btn = QPushButton("📤 Select Target")
        self.select_target_btn.setMaximumWidth(120)
        self.select_target_btn.clicked.connect(self.select_target_action)  # Bind click event
        bututton_layout.addWidget(self.select_target_btn)
        
        # Create remove target box button
        self.remove_target_btn = QPushButton("❌ Remove Target")
        self.remove_target_btn.setMaximumWidth(140)
        self.remove_target_btn.clicked.connect(self.remove_target_action)  # Bind click event
        bututton_layout.addWidget(self.remove_target_btn)
        
        # Create export report button
        self.export_niigz_report_btn = QPushButton("📤 Export Report")
        self.export_niigz_report_btn.setMaximumWidth(120)
        self.export_niigz_report_btn.clicked.connect(self.export_niigz_report)  # Bind click event
        bututton_layout.addWidget(self.export_niigz_report_btn)

        # Create clear results button
        self.clear_niigz_result_btn = QPushButton("🗑️ Clear Results")
        self.clear_niigz_result_btn.setMaximumWidth(120)
        bututton_layout.addWidget(self.clear_niigz_result_btn)
        self.clear_niigz_result_btn.clicked.connect(self.clear_niigz_results)

        # 将按钮布局添加到主布局的右侧
        layout.addLayout(bututton_layout)

        return tab

    def clear_niigz_results(self):
        """Clear NIfTI detection results"""
        self.log_message("🗑️ Clearing NIfTI detection results")
        self.axial_display.setText("Not detected")
        self.sagittal_display.setText("Not detected")
        self.coronal_display.setText("Not detected")
        self.niigz_table.setRowCount(0)
        self.niigz_summary_label.setText("Detection Summary")

    def export_niigz_report(self):
        """Export NIfTI detection report"""
        self.log_message("📤 Exporting NIfTI detection report")
        
        # Check if there are detection results
        if not hasattr(self, 'detection_results') or not self.detection_results:
            self.log_message("❌ No detection results to export")
            QMessageBox.warning(self, "Warning", "No detection results available. Please run detection first.")
            return
        
        # Generate file name with timestamp
        now = datetime.now()
        formatted_time = now.strftime("%Y%m%d_%H%M%S")
        default_filename = f"MRI_Diagnostic_Report_{formatted_time}.pdf"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Report", default_filename, "PDF Files (*.pdf);;All Files (*)"
        )

        if not file_path:
            return

        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            import tempfile
            import os

            # Register font
            font_path = "C:\Windows\Fonts\msyh.ttc"  # Font file path
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                font_name = 'CustomFont'
            else:
                font_name = 'simhei'  # Alternative font

            # Create document
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []

            # Style definitions
            styles = getSampleStyleSheet()

            # Custom title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # Center
                fontName=font_name
            )
            # Custom heading style
            big_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Heading4'],
                fontSize=12,
                spaceAfter=10,
                fontName=font_name
            )

            # Custom paragraph style
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=8,
                spaceAfter=8,
                leading=12,
                fontName=font_name
            )

            # Report title
            title = Paragraph("MRI Examination Report", title_style)
            story.append(title)
            story.append(Spacer(1, 20))

            # Basic information table
            info_data = [
                ["Name:", "", "Gender:", "", "Age:", ""],
                ["Date:", datetime.now().strftime("%Y-%m-%d"), "ID:", "", "Department:", ""],
                ["Site:", "Brain", "Sequence:", "T1/T2", "Model:", "MRI Scanner"]
            ]

            info_table = Table(info_data, colWidths=[60, 90, 60, 60, 80, 80])
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('SPAN', (1, 0), (1, 0)),  # Merge cells
                ('SPAN', (3, 0), (3, 0)),
                ('SPAN', (5, 0), (5, 0)),
            ]))

            story.append(info_table)
            story.append(Spacer(1, 10))

            # Organize results by slice
            slice_results = {}
            for result in self.detection_results:
                slice_name = result['slice']
                if slice_name not in slice_results:
                    slice_results[slice_name] = []
                slice_results[slice_name].append(result)

            # Get images for each slice
            slice_images = {}
            temp_files = []
            
            for slice_name in ['Axial', 'Sagittal', 'Coronal']:
                if slice_name in slice_results and slice_results[slice_name]:
                    # Get the first result for each slice
                    result = slice_results[slice_name][0]
                    img_array = result['result_img']
                    
                    # Save image to temporary file
                    temp_file = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                    temp_files.append(temp_file.name)
                    
                    # Convert numpy array to PIL Image and save
                    from PIL import Image as PILImage
                    import cv2
                    
                    # Convert BGR to RGB if needed
                    if img_array.shape[2] == 3:
                        img_array = cv2.cvtColor(img_array, cv2.COLOR_BGR2RGB)
                    
                    pil_img = PILImage.fromarray(img_array)
                    pil_img.save(temp_file.name)
                    
                    slice_images[slice_name] = temp_file.name

            # Add slice images in a row
            if slice_images:
                story.append(Paragraph("Detection Results :", big_style))
                
                # Create a table for images with spacing
                img_data = []
                img_row = []
                col_widths = []
                
                # Calculate optimal image size while maintaining aspect ratio
                max_img_width = 6 * inch
                max_img_height = 6 * inch
                spacing = 0.2 * inch  # Space between images
                
                from PIL import Image as PILImage
                
                for slice_name, img_path in slice_images.items():
                    # Get original image size
                    with PILImage.open(img_path) as pil_img:
                        orig_width, orig_height = pil_img.size
                    
                    # Calculate aspect ratio
                    aspect_ratio = orig_width / orig_height
                    
                    # Calculate display size while maintaining aspect ratio
                    if aspect_ratio > 1:  # Landscape
                        display_width = min(max_img_width, orig_width * 0.6)
                        display_height = display_width / aspect_ratio
                    else:  # Portrait or square
                        display_height = min(max_img_height, orig_height * 0.6)
                        display_width = display_height * aspect_ratio
                    
                    # Ensure all images have the same height for alignment
                    display_height = min(display_height, max_img_height)
                    display_width = display_height * aspect_ratio
                    
                    # Add image to row
                    img = Image(img_path, width=display_width, height=display_height)
                    img_row.append(img)
                    col_widths.append(display_width)
                    
                    # Add spacing between images (except after last image)
                    if slice_name != list(slice_images.keys())[-1]:
                        img_row.append(Paragraph("", normal_style))
                        col_widths.append(spacing)
                
                if img_row:
                    img_table = Table([img_row], colWidths=col_widths)
                    img_table.setStyle(TableStyle([
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 0),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                        ('TOPPADDING', (0, 0), (-1, -1), 0),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
                    ]))
                    story.append(img_table)
                    story.append(Spacer(1, 10))

            # Calculate measurements for each slice
            slice_measurements = {}
            for slice_name in ['Axial', 'Sagittal', 'Coronal']:
                if slice_name in slice_results and slice_results[slice_name]:
                    # Get the largest tumor in each slice
                    largest_tumor = max(slice_results[slice_name], key=lambda x: x['area'])
                    slice_measurements[slice_name] = {
                        'max_diameter': largest_tumor['diameter'],
                        'area': largest_tumor['area'],
                        'class': largest_tumor['class'],
                        'confidence': largest_tumor['confidence']
                    }

            # Estimate volume based on measurements
            volume_estimate = 0
            if all(slice_name in slice_measurements for slice_name in ['Axial', 'Sagittal', 'Coronal']):
                # Simple volume estimation using the product of diameters
                axial_dia = slice_measurements['Axial']['max_diameter']
                sagittal_dia = slice_measurements['Sagittal']['max_diameter']
                coronal_dia = slice_measurements['Coronal']['max_diameter']
                
                # Assuming ellipsoid shape: V = (4/3) * π * (a/2) * (b/2) * (c/2)
                volume_estimate = (4/3) * 3.14159 * (axial_dia/2) * (sagittal_dia/2) * (coronal_dia/2)

            # Add measurements table
            story.append(Paragraph("Measurements:", big_style))
            
            measurement_data = [['Slice', 'Lesion Type', 'Confidence', 'Max Diameter', 'Area']]
            
            for slice_name in ['Axial', 'Sagittal', 'Coronal']:
                if slice_name in slice_measurements:
                    meas = slice_measurements[slice_name]
                    measurement_data.append([
                        slice_name,
                        meas['class'],
                        f"{meas['confidence']:.2f}",
                        f"{meas['max_diameter']:.2f} mm",
                        f"{meas['area']:.2f} mm²"
                    ])
                else:
                    measurement_data.append([slice_name, "No lesion detected", "-", "-", "-"])

            # Add volume estimate (only if tumor detected)
            if volume_estimate > 0:
                measurement_data.append(['', '', '', '', ''])  # Empty row for spacing
                measurement_data.append(['Volume Estimate:', f"{volume_estimate:.2f} mm³", '', '', ''])
            
            measurement_table = Table(measurement_data, colWidths=[90, 80, 60, 80, 80])
            measurement_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -2), 1, colors.black),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ]))
            
            story.append(measurement_table)
            story.append(Spacer(1, 10))

            # Scan findings
            story.append(Paragraph("Scan Findings:", big_style))
            
            findings_parts = []
            for slice_name in ['Axial', 'Sagittal', 'Coronal']:
                if slice_name in slice_measurements:
                    meas = slice_measurements[slice_name]
                    findings_parts.append(
                        f"{slice_name} slice: {meas['class']} lesion detected with confidence {meas['confidence']:.2f}. Maximum diameter: {meas['max_diameter']:.2f} mm. Area: {meas['area']:.2f} mm²."
                    )
                else:
                    findings_parts.append(
                        f"{slice_name} slice: No obvious abnormalities found."
                    )
            
            if volume_estimate > 0:
                findings_parts.append(f"Estimated lesion volume: {volume_estimate:.2f} mm³.")
            
            findings_text = " ".join(findings_parts)
            findings_para = Paragraph(findings_text, normal_style)
            story.append(findings_para)
            story.append(Spacer(1, 10))

            # Diagnostic conclusion
            story.append(Paragraph("Diagnostic Conclusion:", big_style))
            
            if slice_measurements:
                lesion_types = set([meas['class'] for meas in slice_measurements.values()])
                conclusion_text = f"Multiple lesions detected in brain MRI. {', '.join(lesion_types)} lesion(s) identified with varying confidence levels. Further clinical correlation recommended."
            else:
                conclusion_text = "No obvious abnormalities detected in brain MRI."
            
            conclusion_para = Paragraph(conclusion_text, normal_style)
            story.append(conclusion_para)
            story.append(Spacer(1, 10))

            # Add report footer
            footer_style = ParagraphStyle(
                'CustomFooter',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1,  # Center
                fontName=font_name
            )
            
            footer = Paragraph("This report is generated by AI-assisted diagnostic system. For clinical reference only.", footer_style)
            story.append(Spacer(1, 10))
            story.append(footer)

            # Build PDF
            doc.build(story)
            
            # Clean up temporary files
            for temp_file in temp_files:
                if os.path.exists(temp_file):
                    try:
                        os.unlink(temp_file)
                    except:
                        pass
            
            self.log_message(f"✅ Report exported successfully: {file_path}")
            
            # Create message box with Open button
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowTitle("Success")
            msg_box.setText(f"Report exported successfully to:\n{file_path}")
            msg_box.addButton(QMessageBox.Ok)
            
            # Add Open button
            open_button = msg_box.addButton("Open PDF", QMessageBox.ActionRole)
            
            # Show dialog and get user choice
            msg_box.exec()
            
            # If user clicked Open button, open the PDF file
            if msg_box.clickedButton() == open_button:
                try:
                    import os
                    import sys
                    if sys.platform.startswith('darwin'):  # macOS
                        os.system(f'open "{file_path}"')
                    elif sys.platform.startswith('win'):  # Windows
                        os.startfile(file_path)
                    else:  # Linux/Unix
                        os.system(f'xdg-open "{file_path}"')
                    self.log_message(f"📄 Opening PDF: {file_path}")
                except Exception as e:
                    self.log_message(f"❌ Failed to open PDF: {str(e)}")
                    QMessageBox.warning(self, "Warning", f"Failed to open PDF file:\n{str(e)}")
            
        except Exception as e:
            self.log_message(f"❌ Error exporting report: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to export report:\n{str(e)}")

    def start_niigz_detection(self):
        """Start NIfTI detection"""
        self.log_message("🧠 Starting NIfTI file detection")
        # Get currently selected file
        self.tab_widget.setCurrentIndex(1)
        current_file = self.current_source_path
        if not current_file:
            self.log_message("❌ Error: No NIfTI file selected")
            QMessageBox.warning(self, "Warning", "Please select a NIfTI file first")
            return

        # Check if file is in niigz format
        if not (current_file.endswith('.nii.gz') or current_file.endswith('.nii')):
            self.log_message(f"❌ Error: Unsupported file format: {current_file}")
            QMessageBox.warning(self, "Warning", "Please select a .nii.gz or .nii format file")
            return

        # Check if model is loaded
        if not self.model:
            self.log_message("❌ Error: Model not loaded")
            QMessageBox.warning(self, "Warning", "Please load the model first")
            return

        # Update UI state
        self.start_btn.setEnabled(False)
        self.niigz_summary_label.setText("Detecting...")

        # Start detection thread
        thread = threading.Thread(target=self._process_niigz_file, args=(current_file,))
        thread.daemon = True
        thread.start()


    def _process_niigz_file(self, file_path):
        """Process NIfTI file"""
        self.log_message(f"🧠 Starting to process NIfTI file: {os.path.basename(file_path)}")
        # Get current model path
        model_path = self.model_combo.currentData()
        if not model_path:
            # If no model selected, use default model
            self.log_message("🔍 No model selected, using default model")
            models = self.model_manager.scan_models()
            if models:
                model_path = models[0]['path']
                self.log_message(f"✅ Using default model: {os.path.basename(model_path)}")
            else:
                self.log_message("❌ Error: No model files found")
                self._update_niigz_status("Error: No model files found")
                return

        # Create TumorSliceFinder instance
        output_dir = os.path.join(os.path.dirname(file_path), "niigz_results")
        finder = TumorSliceFinder(
            model_path='pt_models/best.pt',
            nii_path=file_path,
            output_project=output_dir,
            conf=0.65
        )

        # Find optimal slice
        self.log_message("🔍 Finding optimal tumor slice...")
        results = finder.find_best_slices()

        if not results['has_tumor']:
            self.log_message("⚠️  No tumor detected")
            self._update_niigz_status("No tumor detected")
            return

        # Get indices for three planes
        axial_slice = results['axial_slice']
        sagittal_slice = results['sagittal_slice']
        coronal_slice = results['coronal_slice']
        self.log_message(f"✅ Tumor found, slice positions: Axial={axial_slice}, Sagittal={sagittal_slice}, Coronal={coronal_slice}")

        # Load NIfTI data
        self.log_message("📊 Loading NIfTI data...")
        img = nib.load(file_path)
        data = img.get_fdata()
        self.log_message(f"📊 Data shape: {data.shape}, Data type: {data.dtype}")

        # Normalize to uint8
        if data.dtype != np.uint8:
            self.log_message("🔄 Normalizing data to uint8...")
            data = np.clip(data, 0, np.percentile(data, 99))
            data = ((data - data.min()) / (data.max() - data.min()) * 255).astype(np.uint8)

        # Get images for three planes
        axial_img = data[:, :, axial_slice].T
        sagittal_img = data[sagittal_slice, :, :]
        coronal_img = data[:, coronal_slice, :]

        # Use model to detect each slice
        model = self.model
        # class_names = list(model.names.values())

        # Detection results list
        self.detection_results = []

        # Detect axial slice
        axial_results = self._detect_slice(axial_img, "Axial")
        self.detection_results.extend(axial_results)

        # Detect sagittal slice
        sagittal_results = self._detect_slice(sagittal_img, "Sagittal")
        self.detection_results.extend(sagittal_results)

        # Detect coronal slice
        coronal_results = self._detect_slice(coronal_img, "Coronal")
        self.detection_results.extend(coronal_results)

        # Update UI
        self.log_message(f"📋 Detection completed, found {len(self.detection_results)} targets")
        self._update_niigz_ui(axial_img, sagittal_img, coronal_img, self.detection_results, results)

        # Restore UI state
        self.start_btn.setEnabled(True)
        self.log_message("✅ NIfTI file processing completed")

    def _detect_slice(self, slice_img, slice_name):
        """Detect single slice"""
        self.log_message(f"🔍 Detecting {slice_name} slice...")
        results = []
        # Convert 2D slice to 3-channel RGB
        slice_rgb = np.stack([slice_img] * 3, axis=-1)
        # Ensure image data is contiguous in memory
        slice_rgb = np.ascontiguousarray(slice_rgb)
        # Model prediction
        private_path = base_dir / "pt_models" / "Brain_Tumor.pt"
        private_model = YOLO(private_path)
        predictions = private_model.predict(source=slice_rgb, conf=self.conf_slider.value() / 100, verbose=False,show_conf=False)
        for pred in predictions:
            if pred.boxes:
                boxes = pred.boxes.cpu().numpy()
                self.log_message(f"✅ {slice_name} slice detected {len(boxes)} targets")
                for box in boxes:
                    x1, y1, x2, y2 = box.xyxy[0]
                    conf = box.conf[0]
                    cls = int(box.cls[0])
                    class_name = private_model.names[cls] if cls < len(private_model.names) else f"Class{cls}"
                    # Calculate diameter and area (using pixel values for now, actual calculation requires voxel spacing)
                    width = x2 - x1
                    height = y2 - y1
                    diameter = np.sqrt(width**2 + height**2)
                    area = width * height
                    self.log_message(f"📌 {slice_name} slice: {class_name} (Confidence: {conf:.2f}, Coordinates: ({int(x1)}, {int(y1)}, {int(x2)}, {int(y2)}))")
                    results.append({
                        'slice': slice_name,
                        'class': class_name,
                        'confidence': conf,
                        'coordinates': f"({int(x1)}, {int(y1)}, {int(x2)}, {int(y2)})",
                        'diameter': diameter,
                        'area': area,
                        'result_img':predictions[0].plot()

                    })
            else:
                self.log_message(f"⚠️  No targets detected in {slice_name} slice")
        return results

    def _update_niigz_ui(self, axial_img, sagittal_img, coronal_img, detection_results, finder_results):
        """Update NIfTI detection UI"""
        self.log_message("🖥️ Updating NIfTI detection UI...")
        
        # Save original images for later removal of target boxes
        self.original_axial_img = axial_img
        self.original_sagittal_img = sagittal_img
        self.original_coronal_img = coronal_img
        
        # Update display for three slices
        self._update_slice_display(self.axial_display, axial_img)
        self._update_slice_display(self.sagittal_display, sagittal_img)
        self._update_slice_display(self.coronal_display, coronal_img)

        # Update table
        self.niigz_table.setRowCount(0)
        for i, result in enumerate(detection_results):
            self.niigz_table.insertRow(i)
            self.niigz_table.setItem(i, 0, QTableWidgetItem(result['slice']))
            self.niigz_table.setItem(i, 1, QTableWidgetItem(result['class']))
            self.niigz_table.setItem(i, 2, QTableWidgetItem(f"{result['confidence']:.2f}"))
            self.niigz_table.setItem(i, 3, QTableWidgetItem(result['coordinates']))
            self.niigz_table.setItem(i, 4, QTableWidgetItem(f"{result['diameter']:.2f}"))
            self.niigz_table.setItem(i, 5, QTableWidgetItem(f"{result['area']:.2f}"))

        # Update summary
        tumor_center = finder_results['tumor_center']
        summary = f"🧠 Tumor detected | Center: {tumor_center} | Axial: {finder_results['axial_slice']} | Sagittal: {finder_results['sagittal_slice']} | Coronal: {finder_results['coronal_slice']} | Targets: {len(detection_results)}"
        self.niigz_summary_label.setText(summary)
        self.log_message(f"📊 Detection summary: {summary}")

    def _update_slice_display(self, label, img):
        """Update slice display"""
        # Ensure array is C-contiguous
        img_contiguous = np.ascontiguousarray(img)
        # Convert numpy array to QImage
        if len(img_contiguous.shape) == 2:
            # Grayscale image
            height, width = img_contiguous.shape
            bytes_per_line = width
            q_img = QImage(img_contiguous.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
        else:
            # RGB image
            height, width, channels = img_contiguous.shape
            bytes_per_line = channels * width
            q_img = QImage(img_contiguous.data, width, height, bytes_per_line, QImage.Format_RGB888)

        # Scale image to fit label
        scaled_img = q_img.scaled(label.width(), label.height(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        pixmap = QPixmap.fromImage(scaled_img)
        label.setPixmap(pixmap)

    def _update_niigz_status(self, message):
        """Update NIfTI detection status"""
        self.log_message(f"📢 Updating NIfTI detection status: {message}")
        self.niigz_summary_label.setText(message)
        self.start_btn.setEnabled(True)


    def select_target_action(self):
        """Handle select target button click event"""
        if hasattr(self, 'detection_results') and self.detection_results:
            self._update_slice_display(self.axial_display, self.detection_results[0]['result_img'])
            self._update_slice_display(self.sagittal_display, self.detection_results[1]['result_img'])
            self._update_slice_display(self.coronal_display, self.detection_results[2]['result_img'])
            self.log_message(f"✅ Target selected")
        else:
            self.log_message(f"⚠️  No detection results available")

    def remove_target_action(self):
        """Handle remove target button click event - display original images without bounding boxes"""
        if hasattr(self, 'original_axial_img') and hasattr(self, 'original_sagittal_img') and hasattr(self, 'original_coronal_img'):
            self._update_slice_display(self.axial_display, self.original_axial_img)
            self._update_slice_display(self.sagittal_display, self.original_sagittal_img)
            self._update_slice_display(self.coronal_display, self.original_coronal_img)
            self.log_message(f"✅ Target boxes removed, displaying original images")
        else:
            self.log_message(f"⚠️  No original images available")
    def create_nifti_conversion_tab(self):
        """Create NIfTI format conversion tab"""
        nifti_tab = QWidget()
        layout = QVBoxLayout(nifti_tab)

        # File selection area
        file_group = QGroupBox("📁 File Selection")
        file_layout = QVBoxLayout(file_group)

        file_select_layout = QHBoxLayout()
        self.nii_file_edit = QLineEdit()
        self.nii_file_edit.setPlaceholderText("Select NIfTI file or directory...")
        file_select_layout.addWidget(self.nii_file_edit)

        self.browse_nii_btn = QPushButton("Browse")
        self.browse_nii_btn.clicked.connect(self.browse_nii_file)
        file_select_layout.addWidget(self.browse_nii_btn)

        file_layout.addLayout(file_select_layout)

        # Output directory settings
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("📤 Output Directory:"))
        self.output_dir_edit = QLineEdit()
        self.output_dir_edit.setPlaceholderText("Automatically set to input directory + '_swift_normal'")
        output_layout.addWidget(self.output_dir_edit)

        self.browse_output_btn = QPushButton("Browse")
        self.browse_output_btn.clicked.connect(self.browse_output_dir)
        output_layout.addWidget(self.browse_output_btn)

        file_layout.addLayout(output_layout)

        layout.addWidget(file_group)

        # Slice settings and file information horizontal layout area
        settings_layout = QHBoxLayout()
        # Slice settings area
        slice_group = QGroupBox("🔪 Slice Settings")
        slice_layout = QVBoxLayout(slice_group)

        # Slice direction selection
        direction_layout = QHBoxLayout()
        direction_layout.addWidget(QLabel("🧭 Slice Direction:"))
        self.slice_direction_combo = QComboBox()
        self.slice_direction_combo.addItems(["Sagittal", "Coronal", "Axial"])
        self.slice_direction_combo.setCurrentText("Coronal")
        self.slice_direction_combo.currentTextChanged.connect(self.update_slice_info)
        direction_layout.addWidget(self.slice_direction_combo)
        direction_layout.addStretch()

        slice_layout.addLayout(direction_layout)

        # Slice range settings
        range_layout = QHBoxLayout()
        # In the part where slice range settings are created, add signal connection for QSpinBox
        range_layout.addWidget(QLabel("📏 Slice Range:"))
        self.start_slice_spin = QSpinBox()
        self.start_slice_spin.setMinimum(0)
        self.start_slice_spin.setValue(60)
        self.start_slice_spin.valueChanged.connect(self.on_slice_range_changed)  # Add this line
        range_layout.addWidget(self.start_slice_spin)

        range_layout.addWidget(QLabel(" - "))

        self.end_slice_spin = QSpinBox()
        self.end_slice_spin.setMinimum(0)
        self.end_slice_spin.setMaximum(1000)
        self.end_slice_spin.setValue(150)
        self.end_slice_spin.valueChanged.connect(self.on_slice_range_changed)  # Add this line
        range_layout.addWidget(self.end_slice_spin)

        range_layout.addStretch()
        slice_layout.addLayout(range_layout)

        # Slice information display
        info_layout = QHBoxLayout()
        self.slice_info_label = QLabel("_slices: 0, Current range: 0-0")
        info_layout.addWidget(self.slice_info_label)
        info_layout.addStretch()
        slice_layout.addLayout(info_layout)

        # layout.addWidget(slice_group)

        # File information display area
        info_group = QGroupBox("📊 File Information")
        info_layout = QVBoxLayout(info_group)

        self.file_info_text = QTextEdit()
        self.file_info_text.setReadOnly(True)
        self.file_info_text.setMaximumHeight(100)
        info_layout.addWidget(self.file_info_text)

        # layout.addWidget(info_group)
        # Add two areas to horizontal layout
        settings_layout.addWidget(slice_group)
        settings_layout.addWidget(info_group)
        slice_group.setMaximumHeight(150)  # Limit slice settings area height
        info_group.setMaximumHeight(150)  # Limit file information area height
        # Set both areas to equal width
        settings_layout.setStretch(0, 1)
        settings_layout.setStretch(1, 1)

        layout.addLayout(settings_layout)
        # Preview area
        preview_group = QGroupBox("🖼️ Slice Preview")
        preview_layout = QVBoxLayout(preview_group)

        self.preview_scroll = QScrollArea()
        self.preview_scroll.setWidgetResizable(True)
        self.preview_widget = QWidget()
        self.preview_layout = QHBoxLayout(self.preview_widget)
        self.preview_scroll.setWidget(self.preview_widget)

        preview_layout.addWidget(self.preview_scroll)
        layout.addWidget(preview_group)

        # Control buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()  # Push buttons to the right
        
        self.convert_btn = QPushButton("🔄 Convert")
        self.convert_btn.clicked.connect(self.convert_nifti)
        button_layout.addWidget(self.convert_btn)

        self.preview_btn = QPushButton("👀 Preview")
        self.preview_btn.clicked.connect(self.generate_preview)
        button_layout.addWidget(self.preview_btn)

        layout.addLayout(button_layout)

        # Initialize state
        self.current_nii_file = None
        self.nii_data = None

        return nifti_tab

    def on_slice_range_changed(self, value):
        """Update preview when slice range changes"""
        # Set flag indicating slice range has changed
        self.slice_range_changed = True

        # Restart timer to delay preview update
        self.slice_update_timer.start(100)  # 300ms delay to avoid frequent updates

    def browse_nii_file(self):
        """Browse NIfTI file or directory"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select NIfTI File",
            "",
            "NIfTI Files (*.nii *.nii.gz);;All Files (*)"
        )

        if file_path:
            self.nii_file_edit.setText(file_path)
            self.load_nifti_file(file_path)

    def browse_output_dir(self):
        """Browse output directory"""
        dir_path = QFileDialog.getExistingDirectory(self, "Select Output Directory")
        if dir_path:
            self.output_dir_edit.setText(dir_path)

    def load_nifti_file(self, file_path):
        """Load NIfTI file and display information"""
        try:
            import nibabel as nib
            self.current_nii_file = file_path
            nii = nib.load(file_path)
            self.nii_data = nii.get_fdata()

            # Update output directory
            if not self.output_dir_edit.text():
                input_dir = Path(file_path).parent
                output_dir = input_dir / f"{Path(file_path).stem}_swift_normal"
                self.output_dir_edit.setText(str(output_dir))

            # Update slice range
            self.update_slice_range()

            # Display file information
            self.display_file_info(nii)

            # Generate preview
            self.generate_preview()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load NIfTI file: {str(e)}")

    def update_slice_range(self):
        """Update slice range controls"""
        if self.nii_data is not None:
            # Determine maximum number of slices based on slice direction
            direction = self.slice_direction_combo.currentIndex()
            max_slices = self.nii_data.shape[direction] - 1

            self.start_slice_spin.setMaximum(max_slices)
            self.end_slice_spin.setMaximum(max_slices)
            # self.end_slice_spin.setValue(max_slices)

            self.update_slice_info()

    def update_slice_info(self):
        """Update slice information display"""
        if self.nii_data is not None:
            direction = self.slice_direction_combo.currentIndex()
            max_slices = self.nii_data.shape[direction]
            start = self.start_slice_spin.value()
            end = min(self.end_slice_spin.value(), max_slices - 1)

            self.slice_info_label.setText(f"_slices: {max_slices}, Current range: {start}-{end}")

    def display_file_info(self, nii):
        """Display NIfTI file information"""
        try:
            import os
            from datetime import datetime

            file_path = self.current_nii_file
            file_size = os.path.getsize(file_path)
            mod_time = datetime.fromtimestamp(os.path.getmtime(file_path))

            header = nii.header
            shape = nii.shape
            dtype = header.get_data_dtype()
            affine = nii.affine

            # Calculate spatial resolution
            voxel_sizes = header.get_zooms()

            info_text = f"File size: {self.format_file_size(file_size)}\n"
            info_text += f"Modified date: {mod_time.strftime('%Y-%m-%d %H:%M:%S')}\n"
            info_text += f"Image dimensions: {shape}\n"
            info_text += f"Data type: {dtype}\n"
            info_text += f"Spatial resolution: {voxel_sizes[:3] if len(voxel_sizes) >= 3 else voxel_sizes}\n"

            self.file_info_text.setText(info_text)
        except Exception as e:
            self.file_info_text.setText(f"Unable to read file information: {str(e)}")

    def format_file_size(self, size_bytes):
        """Format file size display"""
        if size_bytes == 0:
            return "0 B"

        size_names = ["B", "KB", "MB", "GB"]
        i = 0
        while size_bytes >= 1024 and i < len(size_names) - 1:
            size_bytes /= 1024.0
            i += 1

        return f"{size_bytes:.1f} {size_names[i]}"

    def generate_preview(self):
        """Generate preview images"""
        if self.nii_data is None:
            return

        # Clear existing preview
        for i in reversed(range(self.preview_layout.count())):
            widget = self.preview_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)

        try:
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
            import numpy as np

            direction = self.slice_direction_combo.currentIndex()
            max_slices = self.nii_data.shape[direction]
            start = self.start_slice_spin.value()
            end = min(self.end_slice_spin.value(), max_slices - 1)

            # Select 5 representative slices
            indices = np.linspace(start, end, 5, dtype=int)

            for idx in indices:
                # Extract slice
                if direction == 0:  # Sagittal
                    slice_data = self.nii_data[idx, :, :]
                elif direction == 1:  # Coronal
                    slice_data = self.nii_data[:, idx, :]
                else:  # Axial
                    slice_data = self.nii_data[:, :, idx]

                # Create image
                fig = plt.Figure(figsize=(2, 2), dpi=100)
                ax = fig.add_subplot(111)
                ax.imshow(slice_data, cmap='gray')
                ax.set_title(f"Slice {idx}")
                ax.axis('off')

                canvas = FigureCanvas(fig)
                canvas.setToolTip(f"Slice index: {idx}")

                # Add right-click menu functionality
                canvas.setContextMenuPolicy(Qt.CustomContextMenu)
                canvas.customContextMenuRequested.connect(
                    lambda pos, c=canvas, index=idx, dir_=direction:
                    self.show_slice_context_menu(pos, c, index, dir_)
                )
                self.preview_layout.addWidget(canvas)

            plt.close('all')

        except Exception as e:
            error_label = QLabel(f"Failed to generate preview: {str(e)}")
            self.preview_layout.addWidget(error_label)

    def show_slice_context_menu(self, pos, canvas, slice_index, direction):
        """Show slice context menu"""
        context_menu = QMenu(self)

        # Add zoom action
        zoom_action = QAction("🔍 Zoom In", self)
        zoom_action.triggered.connect(
            lambda: self.show_slice_detail(slice_index, direction)
        )
        context_menu.addAction(zoom_action)

        context_menu.exec(canvas.mapToGlobal(pos))

    def show_slice_detail(self, slice_index, direction):
        """Show slice detail dialog"""
        dialog = SliceDetailDialog(self.nii_data, slice_index, direction, self)
        dialog.exec()

    def update_slice_preview(self):
        """Update slice preview"""
        if not self.slice_range_changed:
            return
        # Reset flag
        self.slice_range_changed = False
        # Update slice information display
        self.update_slice_info()
        # Regenerate preview
        if hasattr(self, 'current_nii_file') and self.current_nii_file:
            self.generate_preview()

    def convert_nifti(self):
        """Execute NIfTI conversion"""
        if not self.current_nii_file:
            QMessageBox.warning(self, "Warning", "Please select a NIfTI file first")
            return

        output_dir = self.output_dir_edit.text()
        if not output_dir:
            QMessageBox.warning(self, "Warning", "Please set output directory")
            return

        try:
            import nibabel as nib
            import numpy as np
            from PIL import Image
            import os

            # Create output directory
            os.makedirs(output_dir, exist_ok=True)

            # Load NIfTI file
            nii = nib.load(self.current_nii_file)
            data = nii.get_fdata()

            direction = self.slice_direction_combo.currentIndex()
            start = self.start_slice_spin.value()
            end = min(self.end_slice_spin.value(), data.shape[direction] - 1)

            # Convert slices
            progress = QProgressDialog("Converting...", "Cancel", 0, end - start + 1, self)
            progress.setWindowModality(Qt.WindowModal)
            progress.show()

            count = 0
            for i in range(start, end + 1):
                if progress.wasCanceled():
                    break

                # Extract slice
                if direction == 0:  # Sagittal
                    slice_data = data[i, :, :]
                elif direction == 1:  # Coronal
                    slice_data = data[:, i, :]
                else:  # Axial
                    slice_data = data[:, :, i]

                # Normalize to 0-255
                slice_data = ((slice_data - slice_data.min()) /
                              (slice_data.max() - slice_data.min()) * 255).astype(np.uint8)

                # Save as PNG
                img = Image.fromarray(slice_data)
                img.save(os.path.join(output_dir, f"slice_{i:04d}.png"))

                count += 1
                progress.setValue(count)

            progress.close()
            QMessageBox.information(self, "Success", f"Conversion completed! Saved {count} slices to:\n{output_dir}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Conversion failed: {str(e)}")

    @property
    def create_realtime_tab(self):
        """Create real-time detection tab"""
        widget = QWidget()
        layout_top = QVBoxLayout(widget)
        layout = QHBoxLayout(widget)

        # Original image display
        original_container = QWidget()
        original_layout = QVBoxLayout(original_container)

        original_title = QLabel("📷 Source")
        original_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin: 0px;")
        original_layout.addWidget(original_title)

        self.original_label = QLabel("Waiting for source...")
        self.original_label.setAlignment(Qt.AlignCenter)
        self.original_label.setMinimumSize(500, 400)
        self.original_label.setStyleSheet("""
            background: black;
            color: white;
            font-weight: bold;
            font-size: 12px;
            padding: 15px;
        """)
        original_layout.addWidget(self.original_label)

        # Result image display
        result_container = QWidget()
        result_layout = QVBoxLayout(result_container)

        result_title = QLabel("🎯 Detection Result")
        result_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #2c3e50; margin: 0px;")
        result_layout.addWidget(result_title)

        self.result_label = QLabel("Waiting for detection result...")
        self.result_label.setAlignment(Qt.AlignCenter)
        self.result_label.setMinimumSize(500, 400)
        self.result_label.setStyleSheet("""
            background: black;
            color: white;
            font-weight: bold;
            font-size: 12px;
            padding: 15px;
        """)
        result_layout.addWidget(self.result_label)

        layout.addWidget(original_container)
        layout.addWidget(result_container)
        layout_top.addLayout(layout)
        # Detection result details
        self.result_detail_widget = DetectionResultWidget()
        layout_top.addWidget(self.result_detail_widget)

        # Create a horizontal layout to place the export report button,紧贴在DetectionResultWidget下方
        button_container = QWidget()
        button_container.setMaximumHeight(40)  # Limit button container height
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(0, 0, 0, 0)  # Reduce top and bottom margins
        button_layout.addStretch()  # Add elastic space to push to the right
        self.export_report_btn = QPushButton("📤 Export Report")
        self.export_report_btn.setMaximumWidth(120)
        self.export_report_btn.clicked.connect(self.export_report)
        button_layout.addWidget(self.export_report_btn)
        
        # Clear button
        self.clear_btn = QPushButton("🗑️ Clear")
        self.clear_btn.setMaximumWidth(100)
        self.clear_btn.clicked.connect(self.clear_display_windows)
        button_layout.addWidget(self.clear_btn)
        
        layout_top.addWidget(button_container)
        return widget

    def export_report(self):
        """Export standardized MRI examination report PDF"""
        if not hasattr(self, 'result_label') or not hasattr(self, 'result_detail_widget'):
            QMessageBox.warning(self, "Warning", "Missing required detection result components")
            return

        # Select save path - automatically generate MRI diagnostic report filename
        now = datetime.now()
        formatted_time = now.strftime("%Y%m%d_%H%M%S")
        default_filename = f"MRI_Diagnostic_Report_{formatted_time}.pdf"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Report", default_filename, "PDF Files (*.pdf);;All Files (*)"
        )

        if not file_path:
            return

        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont

            # Register font
            font_path = "C:\Windows\Fonts\msyh.ttc"  # Font file path
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                font_name = 'CustomFont'
            else:
                font_name = 'simhei'  # Alternative font

            # Create document
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            story = []

            # Style definitions
            styles = getSampleStyleSheet()

            # Custom title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1,  # Center
                fontName=font_name
            )
            # Custom heading style
            big_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Heading4'],
                fontSize=12,
                spaceAfter=10,
                fontName=font_name
            )

            # Custom paragraph style
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=8,
                spaceAfter=8,
                leading=12,
                fontName=font_name
            )

            # Report title
            title = Paragraph("MRI Examination Report", title_style)
            story.append(title)
            story.append(Spacer(1, 20))

            # Basic information table
            info_data = [
                ["Name:", "", "Gender:", "", "Age:", ""],
                ["Date:", datetime.now().strftime("%Y-%m-%d"), "ID:", "", "Department:", ""],
                ["Site:", "Brain", "Sequence:", "T1/T2", "Model:", "MRI Scanner"]
            ]

            info_table = Table(info_data, colWidths=[60, 90, 60, 60, 80, 80])
            info_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('SPAN', (1, 0), (1, 0)),  # Merge cells
                ('SPAN', (3, 0), (3, 0)),
                ('SPAN', (5, 0), (5, 0)),
            ]))

            story.append(info_table)
            story.append(Spacer(1, 10))

            # Detection result image
            if hasattr(self, 'result_label') and self.result_label.pixmap():
                # Save current displayed image
                temp_img_path = "temp_result.png"
                self.result_label.pixmap().save(temp_img_path)

                img = Image(temp_img_path, width=2 * inch, height=2 * inch)
                story.append(Paragraph("Detection Result Image:", big_style))
                story.append(img)
                story.append(Spacer(1, 10))

            # MRI scan findings
            story.append(Paragraph("Scan Findings:",big_style))

            # Get detailed detection results
            findings_text = "No obvious abnormalities found"  # Default value

            # Get detection information from result_detail_widget
            if hasattr(self, 'result_detail_widget'):
                # Assume result_detail_widget is DetectionResultWidget
                if hasattr(self.result_detail_widget, 'result_table'):
                    table = self.result_detail_widget.result_table
                    if table.rowCount() > 0:
                        findings_parts = []
                        for row in range(table.rowCount()):
                            # Read lesion type, show "Unknown" if not available
                            class_name = table.item(row, 0).text() if table.item(row, 0) else "Unknown"
                            # Read confidence and replace with "Possibility" description, show "Unknown" if not available
                            confidence = table.item(row, 1).text() if table.item(row, 1) else "Unknown"
                            # Read size information, show "Unknown" if not available
                            size = table.item(row, 3).text() if table.item(row, 3) else "Unknown"

                            # Initialize default values for diameter and voxel size
                            diameter = "Unknown"
                            voxel_size = "Unknown"

                            # Calculate diameter and voxel size only if size information is valid (compatible with common size formats like "2cm×3cm""20,30", etc.)
                            if size != "Unknown":
                                try:
                                    # Process common size formats (like "2×3""20,30""20 30"), extract numbers
                                    # First replace special symbols with separators, then split into width and height
                                    size_clean = size.replace("cm", "").replace("mm", "").replace("×", ",").replace(" ",
                                                                                                                    ",")
                                    width, height = [float(num.strip()) for num in size_clean.split(",") if num.strip()]
                                    diameter = (width+height)/2  # Diameter is the average of width and height
                                    voxel_size = 0.785*width * height  # Voxel size is width multiplied by height
                                except:
                                    # If size format parsing fails, keep default "Unknown"
                                    pass

                            # Concatenate final text: remove position/size, change confidence to possibility, only keep type, possibility, diameter, voxel f"{voxel_value:.2f}"
                            findings_parts.append(
                                f"{class_name} Possibility: {confidence}, Maximum diameter: {diameter}pixel, Estimated pixel area: {voxel_size:.2f}pixel²"
                            )
                        findings_text = ". ".join(findings_parts)

            findings_para = Paragraph(findings_text, normal_style)
            story.append(findings_para)
            story.append(Spacer(1, 10))

            # Diagnostic conclusion
            story.append(Paragraph("Diagnostic Conclusion:",big_style))
            conclusion_text = "Based on imaging findings, brain tumor lesion is considered. Further examination is recommended for confirmation."  # Example text
            conclusion_para = Paragraph(conclusion_text, normal_style)
            story.append(conclusion_para)
            story.append(Spacer(1, 10))

            # Diagnostic recommendations
            story.append(Paragraph("Diagnostic Recommendations:",big_style))
            recommendation_text = "1. Enhanced scan is recommended to further clarify the nature of the lesion<br/>" \
                                  "2. Comprehensive judgment based on clinical symptoms and laboratory test results<br/>" \
                                  "3. Regular follow-up to observe changes in the lesion"
            recommendation_para = Paragraph(recommendation_text, normal_style)
            story.append(recommendation_para)
            story.append(Spacer(1, 10))

            # Notes
            story.append(Paragraph("Notes:",big_style))
            notes_text = "1. Please conduct comprehensive analysis in combination with clinical symptoms<br/>" \
                         "2. If you feel unwell, please seek medical attention promptly<br/>" \
                         "3. Regular follow-up is recommended to monitor disease changes"
            notes_para = Paragraph(notes_text, normal_style)
            story.append(notes_para)
            story.append(Spacer(1, 10))

            # Signature area
            signature_data = [
                ["Reporting Physician:", "_______________", "Reviewing Physician:", "_______________"],
                ["Report Time:", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "", ""],
                ["Hospital Name:", "XXX Hospital Imaging Department", "", ""]
            ]

            signature_table = Table(signature_data, colWidths=[120, 150, 120, 150])
            signature_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            story.append(signature_table)
            disclaimer_paragraph = Paragraph(
                '<para align=center spaceafter=8>This software is a prototype version and is not designed or intended for use in diagnosis or classification of any medical problems or other medical purposes. The user acknowledges and warrants that the software will not be used for such purposes. This prototype is for medical research purposes only. The software is provided "as is" without any warranty of any kind.</para>',
                styles['Normal']
            )
            disclaimer_paragraph.style.fontSize = 5  # Set smaller font
            disclaimer_paragraph.style.leading = 15  # Line spacing
            story.append(disclaimer_paragraph)
            # Build PDF
            doc.build(story)

            # Clean up temporary file
            if os.path.exists(temp_img_path):
                os.remove(temp_img_path)

            # Replace the original line of code:
            # QMessageBox.information(self, "成功", f"报告已成功导出至: {file_path}")

            # Modified to the following code:
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Success")
            msg_box.setText(f"Report successfully exported to: {file_path}")
            msg_box.setIcon(QMessageBox.Information)

            # Add "Open File" button
            open_button = msg_box.addButton("Open File", QMessageBox.ActionRole)
            msg_box.addButton(QMessageBox.Ok)

            msg_box.exec()

            # If user clicked "Open File" button, execute open operation
            if msg_box.clickedButton() == open_button:
                import subprocess
                import platform

                try:
                    if platform.system() == "Windows":
                        os.startfile(file_path)
                    elif platform.system() == "Darwin":  # macOS
                        subprocess.run(["open", file_path])
                    else:  # Linux
                        subprocess.run(["xdg-open", file_path])
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Unable to open file: {str(e)}")


        except ImportError:
            QMessageBox.critical(self, "Error", "Please install reportlab library: pip install reportlab")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error occurred while exporting report: {str(e)}")

    def create_batch_tab(self):
        """Create batch results tab"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # Control Bar
        control_bar = QHBoxLayout()
        control_bar.addWidget(QLabel("📊 Batch Detection Results:"))
        control_bar.addStretch()

        # Navigation Buttons
        self.prev_result_btn = QPushButton("⬅️ Previous")
        self.prev_result_btn.clicked.connect(self.show_prev_result)
        self.prev_result_btn.setEnabled(False)
        control_bar.addWidget(self.prev_result_btn)

        self.result_index_label = QLabel("0/0")
        self.result_index_label.setStyleSheet("font-weight: bold; margin: 0 10px;")
        control_bar.addWidget(self.result_index_label)

        self.next_result_btn = QPushButton("Next ➡️")
        self.next_result_btn.clicked.connect(self.show_next_result)
        self.next_result_btn.setEnabled(False)
        control_bar.addWidget(self.next_result_btn)

        # Save Button
        self.save_results_btn = QPushButton("💾 Save Results")
        self.save_results_btn.clicked.connect(self.save_batch_results)
        self.save_results_btn.setEnabled(False)
        control_bar.addWidget(self.save_results_btn)

        # Clear Button
        self.clear_results_btn = QPushButton("🗑️ Clear Results")
        self.clear_results_btn.clicked.connect(self.clear_batch_results)
        self.clear_results_btn.setEnabled(False)
        control_bar.addWidget(self.clear_results_btn)

        layout.addLayout(control_bar)

        # Image Display
        image_layout = QHBoxLayout()

        self.batch_original_label = QLabel("Batch Detection: Original Image")
        self.batch_original_label.setAlignment(Qt.AlignCenter)
        self.batch_original_label.setMinimumSize(500, 400)
        self.batch_original_label.setStyleSheet("""
            background: black;
            color: white;
            font-weight: bold;
            font-size: 12px;
            padding: 15px;
        """)

        self.batch_result_label = QLabel("Batch Detection: Result Image")
        self.batch_result_label.setAlignment(Qt.AlignCenter)
        self.batch_result_label.setMinimumSize(500, 400)
        self.batch_result_label.setStyleSheet("""
            background: black;
            color: white;
            font-weight: bold;
            font-size: 12px;
            padding: 15px;
        """)

        image_layout.addWidget(self.batch_original_label)
        image_layout.addWidget(self.batch_result_label)
        layout.addLayout(image_layout)

        # Result Information
        self.batch_info_label = QLabel("📁 Select folder to start batch detection...")
        self.batch_info_label.setWordWrap(True)
        self.batch_info_label.setStyleSheet("""
            background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                stop:0 rgba(236, 240, 241, 0.9), stop:1 rgba(189, 195, 199, 0.9));
            padding: 15px;
            border-radius: 8px;
            font-size: 12px;
            color: #2c3e50;
        """)
        layout.addWidget(self.batch_info_label)

        return widget

    def update_time(self):
        """更新时间显示"""
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.time_label.setText(f"🕒 {current_time}")

    def start_all_cameras(self):
        """启动所有选中的摄像头"""
        selected_cameras = [i for i, cb in enumerate(self.camera_checkboxes) if cb.isChecked()]
        for cam_idx in selected_cameras:
            self.start_camera(cam_idx)

    def pause_all_cameras(self):
        """暂停所有摄像头"""
        for widget in self.video_widgets:
            if widget.isVisible():
                widget.pause_detection()

    def stop_all_cameras(self):
        """停止所有摄像头"""
        for widget in self.video_widgets:
            if widget.isVisible():
                widget.stop_detection()

    def clear_all_cameras(self):
        """清空所有摄像头画面"""
        for widget in self.video_widgets:
            if widget.isVisible():
                widget.clear_frame()

    def monitor_all_cameras(self):
        """将所有摄像头切换为监控模式"""
        for widget in self.video_widgets:
            if widget.isVisible():
                widget.set_monitor_mode(True)

    def start_camera(self, camera_index):
        """启动指定摄像头"""
        if 0 <= camera_index < len(self.video_widgets):
            widget = self.video_widgets[camera_index]
            widget.show()
            widget.start_detection(camera_index + 1)  # 假设摄像头ID从1开始

    def update_stats(self, fall_count=0, normal_count=0):
        """更新统计信息"""
        total = fall_count + normal_count
        self.stats_label.setText(f"📊 Total: {total} | Fall: {fall_count} | Normal: {normal_count}")

    def init_model_combo(self):
        """初始化模型下拉框"""
        self.model_combo.clear()
        models = self.model_manager.scan_models()

        if not models:
            self.model_combo.addItem("No models available")
            self.model_combo.setEnabled(False)
        else:
            self.model_combo.addItems([model['name'] for model in models])
            self.model_combo.setEnabled(True)

    def try_load_default_model(self):
        """尝试加载默认模型"""
        if self.model_combo.count() > 0 and self.model_combo.itemText(0) != "No models available":
            first_model = self.model_combo.itemText(0)
            self.load_model_by_name(first_model)

    def load_model_by_name(self, model_name):
        """根据名称加载模型"""
        models = self.model_manager.scan_models()
        for model in models:
            if model['name'] == model_name:
                self.load_model(model['path'])
                break

    def load_model(self, model_path):
        """Load model"""
        try:
            self.model = YOLO(model_path)
            self.log_message(f"✅ Model loaded successfully: {Path(model_path).name}")
            # self.update_button_states()
            return True
        except Exception as e:
            self.log_message(f"❌ Model loading failed: {str(e)}")
            self.model = None
            return False

    def show_model_selection_dialog(self):
        """Show model selection dialog"""
        dialog = ModelSelectionDialog(self.model_manager, self)
        if dialog.exec() == QDialog.Accepted and dialog.selected_model:
            if self.load_model(dialog.selected_model):
                model_name = Path(dialog.selected_model).name
                # 更新下拉框
                index = self.model_combo.findText(model_name)
                if index >= 0:
                    self.model_combo.setCurrentIndex(index)
                else:
                    self.model_combo.addItem(model_name)
                    self.model_combo.setCurrentText(model_name)

    def refresh_camera_list(self):
        """刷新摄像头列表"""
        self.camera_manager.scan_cameras()
        self.camera_combo.clear()

        cameras = self.camera_manager.get_available_cameras()
        if cameras:
            for camera in cameras:
                self.camera_combo.addItem(f"{camera['name']} ({camera['resolution']})", camera['id'])
        else:
            self.camera_combo.addItem("No cameras detected", -1)

    def on_model_changed(self, model_text):
        """Model selection changed"""
        if model_text != "No models available":
            self.load_model_by_name(model_text)

    def on_confidence_changed(self, value):
        """Confidence slider changed"""
        conf_value = value / 100.0
        self.confidence_threshold = conf_value
        self.conf_spinbox.blockSignals(True)
        self.conf_spinbox.setValue(conf_value)
        self.conf_spinbox.blockSignals(False)

    def on_confidence_spinbox_changed(self, value):
        """Confidence spinbox changed"""
        self.confidence_threshold = value
        self.conf_slider.blockSignals(True)
        self.conf_slider.setValue(int(value * 100))
        self.conf_slider.blockSignals(False)

    def on_source_changed(self, source_text):
        """检测源改变"""
        source_map = {
            "📷 Single Image": "image",
            "🎬 Video File": "video",
            "📹 Camera": "camera",
            "📂 Batch Folder": "batch",
            "🧠 NIfTI File": "nifti"
        }
        self.current_source_type = source_map.get(source_text)

        # 显示/隐藏摄像头选择
        is_camera = self.current_source_type == "camera"
        for i in range(self.camera_select_layout.count()):
            item = self.camera_select_layout.itemAt(i)
            if item.widget():
                item.widget().setVisible(is_camera)

        self.current_source_path = None
        self.current_file_label.setText("No file selected yet")
        self.clear_display_windows()
        self.update_button_states()

    def update_button_states(self):
        """更新按钮状态"""
        has_model = self.model is not None

        if self.current_source_type == "camera":
            has_source = self.camera_combo.currentData() != -1
            self.select_file_btn.setEnabled(False)
        else:
            has_source = self.current_source_path is not None
            self.select_file_btn.setEnabled(True)

        self.start_btn.setEnabled(has_model and has_source)

    def select_file(self):
        """Select file or folder"""
        if self.current_source_type == "image":
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Image", "",
                "Image Files (*.jpg *.jpeg *.png *.bmp *.tiff *.webp);;All Files (*)"
            )
        elif self.current_source_type == "video":
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select Video", "",
                "Video Files (*.mp4 *.avi *.mov *.mkv *.wmv *.flv);;All Files (*)"
            )
        elif self.current_source_type == "batch":
            file_path = QFileDialog.getExistingDirectory(self, "Select Folder Containing Images")
        elif self.current_source_type == "nifti":
            file_path, _ = QFileDialog.getOpenFileName(
                self, "Select NIfTI File", "",
                "NIfTI Files (*.nii.gz *.nii);;All Files (*)"
            )
        else:
            return

        if file_path:
            self.current_source_path = file_path
            self.current_file_label.setText(f"📁 Selected: {Path(file_path).name}")
            self.log_message(f"📁 Selected: {file_path}")
            self.update_button_states()

            if self.current_source_type in ["image", "video"]:
                self.preview_file(file_path)

    def preview_file(self, file_path):
        """Preview file"""
        try:
            if self.current_source_type == "image":
                img = cv2.imread(file_path)
                if img is not None:
                    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    self.display_image(img_rgb, self.original_label)
                    self.result_label.clear()
                    self.result_label.setText("Waiting for detection result...")
        except Exception as e:
            self.log_message(f"❌ Failed to preview file: {str(e)}")

    def start_detection(self):
        """Start detection"""
        if not self.model:
            self.log_message("❌ Error: Model not loaded")
            return

        if self.current_source_type == "batch":
            self.start_batch_detection()
        elif self.current_source_type == "nifti":
            self.start_niigz_detection()
        else:
            self.start_single_detection()

    def start_single_detection(self):
        """Start single detection"""
        camera_id = 0
        if self.current_source_type == "camera":
            camera_id = self.camera_combo.currentData()
            if camera_id == -1:
                self.log_message("❌ Error: No available camera")
                return

        self.detection_thread = DetectionThread(
            self.model, self.current_source_type, self.current_source_path, camera_id, self.confidence_threshold
        )
        self.detection_thread.result_ready.connect(self.on_detection_result)
        self.detection_thread.progress_updated.connect(self.progress_bar.setValue)
        self.detection_thread.status_changed.connect(self.statusBar().showMessage)
        self.detection_thread.error_occurred.connect(self.log_message)
        self.detection_thread.finished.connect(self.on_detection_finished)

        self.update_detection_ui_state(True)
        self.tab_widget.setCurrentIndex(0)  # 切换到实时检测

        self.detection_thread.start()
        self.log_message(f"🚀 Starting {self.current_source_type} detection...")

    def start_batch_detection(self):
        """Start batch detection"""
        self.batch_results.clear()

        self.batch_detection_thread = BatchDetectionThread(
            self.model, self.current_source_path, self.confidence_threshold
        )
        self.batch_detection_thread.result_ready.connect(self.on_batch_result)
        self.batch_detection_thread.progress_updated.connect(self.progress_bar.setValue)
        self.batch_detection_thread.current_file_changed.connect(self.statusBar().showMessage)
        self.batch_detection_thread.finished.connect(self.on_batch_finished)

        self.update_detection_ui_state(True)
        self.tab_widget.setCurrentIndex(2)

        self.batch_detection_thread.start()
        self.log_message("🚀 Starting batch detection...")

    def update_detection_ui_state(self, detecting):
        """Update detection UI state"""
        self.start_btn.setEnabled(not detecting)
        self.pause_btn.setEnabled(detecting and self.current_source_type != "batch")
        self.stop_btn.setEnabled(detecting)
        self.source_combo.setEnabled(not detecting)
        self.select_file_btn.setEnabled(not detecting and self.current_source_type != "camera")
        self.model_combo.setEnabled(not detecting)
        # Update snapshot button state
        self.kuaizhao_btn.setEnabled(detecting and self.current_source_type in ["camera", "video"])

    def pause_detection(self):
        """Pause/resume detection"""
        if self.detection_thread and self.detection_thread.is_running:
            if self.detection_thread.is_paused:
                self.detection_thread.resume()
                self.pause_btn.setText("⏸️ Pause")
                self.log_message("▶️ Detection resumed")
            else:
                self.detection_thread.pause()
                self.pause_btn.setText("▶️ Resume")
                self.log_message("⏸️ Detection paused")

    def stop_detection(self):
        """Stop detection"""
        if self.detection_thread and self.detection_thread.is_running:
            self.detection_thread.stop()
            self.detection_thread.wait()

        if self.batch_detection_thread and self.batch_detection_thread.is_running:
            self.batch_detection_thread.stop()
            self.batch_detection_thread.wait()

        self.on_detection_finished()

    def kuaizhao_detection(self):
        """Toggle auto-save monitoring snapshot status"""
        if not self.video_is_auto_saving:
            self.start_auto_save()
        else:
            self.stop_auto_save()

    def start_auto_save(self):
        """Start auto-saving snapshots"""
        if not self.model:
            QMessageBox.warning(self, "Warning", "Please select a model first")
            return

        # Initialize video recorder
        source_name = "Camera" if self.current_source_type == "camera" else "Video"
        if self.current_source_type == "camera":
            source_id = self.camera_combo.currentData()
            source_name = f"Camera{source_id}"
        elif self.current_source_type == "video":
            source_name = Path(self.current_source_path).stem

        self.video_recorder = DetectionVideoRecorder(
            source_name, self.history_dir
        )
        self.video_recorder.start_recording()

        self.video_is_auto_saving = True
        self.kuaizhao_btn.setText("⏹️ Stop Snapshot")
        self.log_message("🎬 Started recording snapshots")

    def stop_auto_save(self):
        """Stop auto-saving snapshots"""
        if self.video_recorder:
            self.video_recorder.stop_recording()
            self.video_recorder = None

        self.video_is_auto_saving = False
        self.kuaizhao_btn.setText("📸 Snapshot")
        self.log_message("⏹️ Stopped recording snapshots")

    def on_detection_result(self, original_img, result_img, inference_time, results, class_names):
        """Detection result callback"""
        # Display images
        self.display_image(original_img, self.original_label)
        self.display_image(result_img, self.result_label)

        # Update result details
        self.result_detail_widget.update_results(results, class_names, inference_time)

        # If recording snapshots, add frame
        if self.video_is_auto_saving and self.video_recorder:
            detection_info = {
                'results': results,
                'class_names': class_names,
                'inference_time': inference_time
            }

            self.video_recorder.add_frame(result_img, detection_info)

        # Record log (simplified version to avoid excessive output)
        if results and results[0].boxes and len(results[0].boxes) > 0:
            object_count = len(results[0].boxes)

            # Count classes
            classes = results[0].boxes.cls.cpu().numpy().astype(int)
            class_counts = {}
            for cls in classes:
                class_name = class_names[cls] if cls < len(class_names) else f"Class{cls}"
                class_counts[class_name] = class_counts.get(class_name, 0) + 1

            class_summary = ", ".join([f"{name}:{count}" for name, count in class_counts.items()])
            self.log_message(f"🎯 Detected {object_count} objects: {class_summary} (Time: {inference_time:.3f}s)")
        else:
            self.log_message(f"⚪ No objects detected (Time: {inference_time:.3f}s)")

    def on_batch_result(self, file_path, original_img, result_img, inference_time, results, class_names):
        """Batch detection result callback"""
        # Calculate object count
        object_count = len(results[0].boxes) if results and results[0].boxes else 0

        # Calculate diameter and voxel information
        diameters = []
        voxels = []
        if results and results[0].boxes:
            boxes = results[0].boxes.xyxy.cpu().numpy()
            for box in boxes:
                # Calculate width and height
                width = box[2] - box[0]
                height = box[3] - box[1]

                # Calculate diameter (assuming circular tumor, diameter equals average size)
                diameter = (width + height) / 2  # in pixels
                diameters.append(diameter)

                # Calculate voxel (simplified as 0.785 * w * h * voxel unit, default voxel unit is 1mm)
                voxel_size = 1.0  # default voxel unit is 1mm
                voxel_value = 0.785 * width * height * voxel_size
                voxels.append(voxel_value)

        # Save result
        result_data = {
            'file_path': file_path,
            'original_img': original_img,
            'result_img': result_img,
            'inference_time': inference_time,
            'results': results,
            'class_names': class_names,
            'object_count': object_count,
            'diameters': diameters,
            'voxels': voxels
        }

        self.batch_results.append(result_data)

        # Display first result
        if len(self.batch_results) == 1:
            self.current_batch_index = 0
            self.show_batch_result(0)

        self.update_batch_navigation()

        # Record log
        filename = Path(file_path).name
        if object_count > 0:
            self.log_message(f"✅ {filename}: {object_count} objects ({inference_time:.3f}s)")
        else:
            self.log_message(f"⚪ {filename}: No objects ({inference_time:.3f}s)")

    def on_batch_finished(self):
        """Batch detection finished"""
        total_count = len(self.batch_results)
        total_objects = sum(result['object_count'] for result in self.batch_results)

        self.log_message(f"🎉 Batch detection completed! Processed {total_count} images, detected {total_objects} objects")
        self.statusBar().showMessage(f"Batch detection completed - {total_count} images, {total_objects} objects")

        self.save_results_btn.setEnabled(True)
        self.clear_results_btn.setEnabled(True)
        self.result_index_label.setText(f"1/{len(self.batch_results)}")
        self.on_detection_finished()

    def on_detection_finished(self):
        """Detection finished callback"""
        self.update_detection_ui_state(False)
        self.pause_btn.setText("⏸️ Pause")
        self.progress_bar.setValue(0)

        # Stop snapshot recording
        if self.video_is_auto_saving:
            self.stop_auto_save()

    def show_batch_result(self, index):
        """Show batch result"""
        if 0 <= index < len(self.batch_results):
            result = self.batch_results[index]

            self.display_image(result['original_img'], self.batch_original_label)
            self.display_image(result['result_img'], self.batch_result_label)

            filename = Path(result['file_path']).name
            object_count = result['object_count']
            inference_time = result['inference_time']

            info_text = f"📁 File: {filename}\n"
            info_text += f"🎯 Detected objects: {object_count}\n"
            info_text += f"⏱️ Inference time: {inference_time:.3f} seconds\n"

            if result['results'] and result['results'][0].boxes and len(result['results'][0].boxes) > 0:
                # Display class statistics
                classes = result['results'][0].boxes.cls.cpu().numpy().astype(int)
                confidences = result['results'][0].boxes.conf.cpu().numpy()

                class_counts = {}
                for cls in classes:
                    class_name = result['class_names'][cls] if cls < len(result['class_names']) else f"Class{cls}"
                    class_counts[class_name] = class_counts.get(class_name, 0) + 1

                info_text += "📊 Class statistics: " + ", ".join(
                    [f"{name}:{count}" for name, count in class_counts.items()]) + ""

                # Determine display method based on number of detected objects
                if len(result['results'][0].boxes) == 1:
                    # Single object: display specific confidence
                    info_text += f"🎯 Confidence: {confidences[0]:.3f}"
                else:
                    # Multiple objects: display average confidence and confidence range
                    info_text += f"🎯 Average confidence: {np.mean(confidences):.3f}"
                    info_text += f"\n📈 Confidence range: {np.min(confidences):.3f} - {np.max(confidences):.3f}"

                # Add diameter and voxel information
                if 'diameters' in result and 'voxels' in result:
                    if len(result['results'][0].boxes) == 1:
                        # Single object: display specific diameter(mm) and voxel(mm²)
                        diameter = result['diameters'][0] if result['diameters'] else 0
                        voxel = result['voxels'][0] if result['voxels'] else 0

                        # Assume 1:1 pixel to mm conversion ratio (calibration may be needed in actual application)
                        diameter_mm = diameter  # Simplified processing here, actual application needs conversion based on camera calibration or known reference object

                        info_text += f"\n📏 Estimated diameter: {diameter_mm:.2f} pixel"
                        info_text += f"\n🧩 Estimated pixel area: {voxel:.2f} pixel²"
                    else:
                        # Multiple objects: display average diameter and voxel, and information for each object
                        avg_diameter = np.mean(result['diameters']) if result['diameters'] else 0
                        avg_voxel = np.mean(result['voxels']) if result['voxels'] else 0

                        # Assume 1:1 pixel to mm conversion ratio
                        avg_diameter_mm = avg_diameter

                        info_text += f"\n📏 Average diameter: {avg_diameter_mm:.2f} pixel"
                        info_text += f"\n🧩 Average pixel: {avg_voxel:.2f} pixel²"

                        info_text += "\n📏 Each object diameter (pixel): " + ", ".join([f"{d:.2f}" for d in result['diameters']])
                        info_text += "\n🧩 Each pixel (pixel²): " + ", ".join([f"{v:.2f}" for v in result['voxels']])

            self.batch_info_label.setText(info_text)
            self.result_index_label.setText(f"{index + 1}/{len(self.batch_results)}")

    def show_prev_result(self):
        """Show previous result"""
        if self.current_batch_index > 0:
            self.current_batch_index -= 1
            self.show_batch_result(self.current_batch_index)
            self.update_batch_navigation()

    def show_next_result(self):
        """Show next result"""
        if self.current_batch_index < len(self.batch_results) - 1:
            self.current_batch_index += 1
            self.show_batch_result(self.current_batch_index)
            self.update_batch_navigation()

    def update_batch_navigation(self):
        """Update batch result navigation"""
        has_results = len(self.batch_results) > 0
        self.prev_result_btn.setEnabled(has_results and self.current_batch_index > 0)
        self.next_result_btn.setEnabled(has_results and self.current_batch_index < len(self.batch_results) - 1)

    def clear_batch_results(self):
        self.batch_results.clear()
        self.batch_result_label.setText('🎯 Batch detection: Result image')
        self.batch_original_label.setText('📷 Batch detection: Original image')
        self.batch_info_label.setText('📁 Select folder to start batch detection...')
        self.result_index_label.setText("0/0")
        self.save_results_btn.setEnabled(False)
        self.next_result_btn.setEnabled(False)
        self.prev_result_btn.setEnabled(False)
        self.clear_results_btn.setEnabled(False)

    def save_batch_results(self):
        """Save batch detection results"""
        if not self.batch_results:
            QMessageBox.information(self, "Information", "No results to save")
            return

        save_dir = QFileDialog.getExistingDirectory(self, "Select Save Directory")
        if not save_dir:
            return

        try:
            save_path = Path(save_dir)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            result_dir = save_path / f"detection_results_{timestamp}"
            result_dir.mkdir(exist_ok=True)

            # Save detection result images
            for i, result in enumerate(self.batch_results):
                file_name = Path(result['file_path']).stem
                result_img = cv2.cvtColor(result['result_img'], cv2.COLOR_RGB2BGR)
                result_save_path = result_dir / f"{file_name}_result.jpg"
                cv2.imwrite(str(result_save_path), result_img)

            # Save detection report (TXT format)
            self.save_detection_report(result_dir)
            
            # Save detection report (Excel format)
            self.save_detection_report_xlsx(result_dir)

            QMessageBox.information(self, "Success", f"Results saved to:\n{result_dir}")
            self.log_message(f"💾 Results saved to: {result_dir}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Save failed: {str(e)}")
            self.log_message(f"❌ Save failed: {str(e)}")

    def save_detection_report(self, result_dir):
        """Save detection report"""
        report_path = result_dir / "detection_report.txt"

        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("🎯 MediScreen-Brain Tumor Detection System - Batch Detection Report\n")
            f.write("=" * 60 + "\n")
            f.write(f"📅 Processing time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"🎚️ Confidence threshold: {self.confidence_threshold}\n")
            f.write(f"📂 Number of images processed: {len(self.batch_results)}\n")
            f.write(f"🎯 Total detected objects: {sum(r['object_count'] for r in self.batch_results)}\n")
            f.write("\n📊 Detailed results:\n")
            f.write("-" * 60 + "\n")

            for i, result in enumerate(self.batch_results, 1):
                f.write(f"{i}. 📁 {Path(result['file_path']).name}\n")
                f.write(f"   🎯 Detected objects: {result['object_count']}\n")
                f.write(f"   ⏱️ Inference time: {result['inference_time']:.3f} seconds\n")

                if result['results'] and result['results'][0].boxes and len(result['results'][0].boxes) > 0:
                    confidences = result['results'][0].boxes.conf.cpu().numpy()
                    classes = result['results'][0].boxes.cls.cpu().numpy().astype(int)

                    # Determine display method based on number of detected objects
                    if len(result['results'][0].boxes) == 1:

                        f.write(f"   📈 Confidence: {confidences[0]:.3f}\n")
                    else:
                        # Multiple objects: display confidence range
                        f.write(f"   📈 Confidence range: {np.min(confidences):.3f} - {np.max(confidences):.3f}\n")

                    # Class statistics
                    class_counts = {}
                    for cls in classes:
                        class_name = result['class_names'][cls] if cls < len(result['class_names']) else f"Class{cls}"
                        class_counts[class_name] = class_counts.get(class_name, 0) + 1

                    f.write("   📊 Class distribution: " + ", ".join(
                        [f"{name}:{count}" for name, count in class_counts.items()]) + "\n")

                    # Add diameter and voxel information
                    if 'diameters' in result and 'voxels' in result:
                        if len(result['results'][0].boxes) == 1:

                            diameter = result['diameters'][0] if result['diameters'] else 0
                            voxel = result['voxels'][0] if result['voxels'] else 0

                            # Assume 1:1 pixel to mm conversion ratio (calibration may be needed in actual application)
                            diameter_mm = diameter  # Simplified processing here, actual application needs conversion based on camera calibration or known reference object

                            f.write(f"   📏 Diameter: {diameter_mm:.2f} pixel\n")
                            f.write(f"   🧩 Pixel: {voxel:.2f} pixel²\n")
                        else:

                            f.write("   📏 Each object diameter (pixel): ")
                            for i, diameter in enumerate(result['diameters']):
                                # Assume 1:1 pixel to mm conversion ratio
                                diameter_mm = diameter
                                if i == len(result['diameters']) - 1:
                                    f.write(f"{diameter_mm:.2f}")
                                else:
                                    f.write(f"{diameter_mm:.2f}, ")
                            f.write("\n")

                            f.write("   🧩 Each pixel (pixel²): ")
                            for i, voxel in enumerate(result['voxels']):
                                if i == len(result['voxels']) - 1:
                                    f.write(f"{voxel:.2f}")
                                else:
                                    f.write(f"{voxel:.2f}, ")
                            f.write("\n")

                f.write("\n")

    def save_detection_report_xlsx(self, result_dir):
        """Save detection report as Excel file"""
        excel_path = result_dir / "detection_report.xlsx"
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detection Results"
        
        # Define styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        
        # Define border
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set column headers
        headers = [
            "Index",
            "File Name",
            "Detected Objects",
            "Inference Time (s)",
            "Confidence",
            "Class Distribution",
            "Max Diameter (pixel)",
            "Total Area (pixel²)"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Adjust column widths
        column_widths = [8, 35, 12, 15, 20, 30, 18, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Write data rows
        for row_idx, result in enumerate(self.batch_results, 2):
            # 序号
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
            
            # 文件名
            file_name = Path(result['file_path']).name
            ws.cell(row=row_idx, column=2, value=file_name).border = thin_border
            
            # 检测目标数
            ws.cell(row=row_idx, column=3, value=result['object_count']).border = thin_border
            
            # 推理时间
            ws.cell(row=row_idx, column=4, value=round(result['inference_time'], 3)).border = thin_border
            
            # 置信度
            if result['results'] and result['results'][0].boxes and len(result['results'][0].boxes) > 0:
                confidences = result['results'][0].boxes.conf.cpu().numpy()
                classes = result['results'][0].boxes.cls.cpu().numpy().astype(int)
                
                if len(confidences) == 1:
                    confidence_str = f"{confidences[0]:.3f}"
                else:
                    confidence_str = f"{np.min(confidences):.3f} - {np.max(confidences):.3f}"
                
                ws.cell(row=row_idx, column=5, value=confidence_str).border = thin_border
                
                # 类别分布
                class_counts = {}
                for cls in classes:
                    class_name = result['class_names'][cls] if cls < len(result['class_names']) else f"Class{cls}"
                    class_counts[class_name] = class_counts.get(class_name, 0) + 1
                
                class_dist_str = ", ".join([f"{name}:{count}" for name, count in class_counts.items()])
                ws.cell(row=row_idx, column=6, value=class_dist_str).border = thin_border
                
                # 直径和面积信息
                if 'diameters' in result and 'voxels' in result:
                    max_diameter = max(result['diameters']) if result['diameters'] else 0
                    total_voxel = sum(result['voxels']) if result['voxels'] else 0
                    
                    ws.cell(row=row_idx, column=7, value=round(max_diameter, 2)).border = thin_border
                    ws.cell(row=row_idx, column=8, value=round(total_voxel, 2)).border = thin_border
                else:
                    ws.cell(row=row_idx, column=7, value="N/A").border = thin_border
                    ws.cell(row=row_idx, column=8, value="N/A").border = thin_border
            else:
                ws.cell(row=row_idx, column=5, value="No detection result").border = thin_border
                ws.cell(row=row_idx, column=6, value="-").border = thin_border
                ws.cell(row=row_idx, column=7, value="N/A").border = thin_border
                ws.cell(row=row_idx, column=8, value="N/A").border = thin_border
        
        # Add summary sheet
        summary_ws = wb.create_sheet(title="Summary")
        
        # Summary headers
        summary_headers = [
            "Metric",
            "Value"
        ]
        
        for col, header in enumerate(summary_headers, 1):
            cell = summary_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        summary_ws.column_dimensions['A'].width = 25
        summary_ws.column_dimensions['B'].width = 20
        
        # Write summary data
        summary_data = [
            ("Processing Time", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Confidence Threshold", self.confidence_threshold),
            ("Number of Images", len(self.batch_results)),
            ("Total Detected Objects", sum(r['object_count'] for r in self.batch_results)),
            ("Average Inference Time (s)", round(sum(r['inference_time'] for r in self.batch_results) / len(self.batch_results), 3) if self.batch_results else 0)
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 2):
            summary_ws.cell(row=row_idx, column=1, value=label).border = thin_border
            summary_ws.cell(row=row_idx, column=2, value=value).border = thin_border
        
        # Save workbook
        wb.save(excel_path)

    def clear_display_windows(self):
        """Clear display windows"""
        self.original_label.clear()
        self.original_label.setText("Waiting for source...")
        self.result_label.clear()
        self.result_label.setText("Waiting for detection result...")
        
        # Clear Detection Result Details
        if hasattr(self, 'result_detail_widget'):
            # Clear the result table
            if hasattr(self.result_detail_widget, 'result_table'):
                self.result_detail_widget.result_table.setRowCount(0)
            # Clear the statistics label
            if hasattr(self.result_detail_widget, 'stats_label'):
                self.result_detail_widget.stats_label.setText("Detection summary...")

    def display_image(self, img_array, label):
        """Display image"""
        if img_array is None:
            return

        height, width, channel = img_array.shape
        bytes_per_line = 3 * width
        q_image = QImage(img_array.data, width, height, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_image)
        scaled_pixmap = pixmap.scaled(label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        label.setPixmap(scaled_pixmap)

    def clear_display(self, lable):
        pass

    def log_message(self, message):
        """Add log message"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{timestamp}] {message}")

        # Limit log lines
        max_lines = 1000
        lines = self.log_text.toPlainText().split('\n')
        if len(lines) > max_lines:
            keep_lines = lines[-500:]
            self.log_text.setPlainText('\n'.join(keep_lines))

        # Auto scroll to bottom
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def clear_log(self):
        """Clear log"""
        self.log_text.clear()
        self.log_message("🗑️ Log cleared")




class DetectionVideoRecorder:
    """Detection video recorder for recording real-time detection snapshots"""

    def __init__(self, source_name, output_dir, fps=20):
        self.source_name = source_name
        self.output_dir = output_dir
        self.fps = fps
        self.is_recording = False
        self.video_writer = None
        self.frames = []
        self.detection_stats = {}
        self.total_detections = 0
        self.start_time = None
        self.end_time = None
        self.max_frames_per_file = fps * 60 * 60 * 24  # 24 hours of video

    def start_recording(self):
        """Start recording"""
        if self.is_recording:
            return

        self.is_recording = True
        self.start_time = time.time()
        self.frames.clear()
        self.detection_stats.clear()
        self.total_detections = 0

        # Generate filename
        timestamp = int(self.start_time)
        self.filename_base = f"{self.source_name}_{timestamp}"
        self.mp4_path = self.output_dir / f"{self.filename_base}.mp4"
        self.json_path = self.output_dir / f"{self.filename_base}.json"

        # Initialize video writer (set later when adding first frame)
        self.video_writer = None

    def add_frame(self, frame, detection_info):
        """Add frame"""
        if not self.is_recording:
            return
        # Check if there are detection results
        if not detection_info or not detection_info.get('results'):
            return

        results = detection_info['results']
        if not hasattr(results[0], 'boxes') or not results[0].boxes or len(results[0].boxes) == 0:
            return
        # If it's the first frame, initialize video writer
        if self.video_writer is None:
            height, width = frame.shape[:2]
            fourcc = cv2.VideoWriter_fourcc(*'mp4v')
            self.video_writer = cv2.VideoWriter(str(self.mp4_path), fourcc, self.fps, (width, height))

        # Write frame - Fix color difference: convert RGB to BGR
        if frame.shape[2] == 3:  # Ensure it's a 3-channel color image
            bgr_frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
            self.video_writer.write(bgr_frame)
        else:
            self.video_writer.write(frame)

        self.frames.append(frame.copy())

        # Update detection statistics
        if detection_info and detection_info.get('results'):
            results = detection_info['results']
            if hasattr(results[0], 'boxes') and results[0].boxes and len(results[0].boxes) > 0:
                self.total_detections += len(results[0].boxes)

                # Count classes
                if hasattr(results[0].boxes, 'cls'):
                    classes = results[0].boxes.cls.cpu().numpy().astype(int)
                    class_names = detection_info.get('class_names', [])

                    for cls in classes:
                        if cls < len(class_names):
                            class_name = class_names[cls]
                            self.detection_stats[class_name] = self.detection_stats.get(class_name, 0) + 1

        # Check if need to save file
        if len(self.frames) >= self.max_frames_per_file:
            self.save_recording()
            self.start_recording()  # Start new recording

    def stop_recording(self):
        """Stop recording"""
        if not self.is_recording:
            return

        self.is_recording = False
        self.end_time = time.time()

        if self.video_writer:
            self.video_writer.release()
            self.video_writer = None

        # Save recording
        if self.frames:
            self.save_recording()

    def save_recording(self):
        """Save recording"""
        if not self.frames or not self.start_time:
            return

        # Ensure video writer is released
        if self.video_writer:
            self.video_writer.release()
            self.video_writer = None

        # Save JSON metadata
        metadata = {
            'camera_id': self.source_name,
            'source_name': self.source_name,
            'start_time': self.start_time,
            'end_time': self.end_time or time.time(),
            'fps': self.fps,
            'total_detections': self.total_detections,
            'detection_stats': self.detection_stats,
            'frame_count': len(self.frames),
            'mp4_filename': self.mp4_path.name,
            'json_filename': self.json_path.name
        }

        with open(self.json_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

        print(f"Saved detection snapshot: {self.source_name} - {len(self.frames)} frames, {self.total_detections} detections")
        print(f"File path: {self.mp4_path}")
        print(f"JSON path: {self.json_path}")


def main():
    app = QApplication(sys.argv)

    # Set application information
    app.setApplicationName("YOLO-based Brain Tumor Detection System")
    app.setApplicationVersion("2.0")
    app.setOrganizationName("[Neuroengineering Lab, UESTC](https://www.neuro.uestc.edu.cn/)")

    # Set high DPI scaling
    # app.setAttribute(Qt.AA_EnableHighDpiScaling)
    # app.setAttribute(Qt.AA_UseHighDpiPixmaps)

    # Create main window
    window = EnhancedDetectionUI()
    window.show()

    # Startup messages
    window.log_message("🚀 MediScreen-Brain Tumor Detection System started")
    window.log_message("✨ New features: Gradient UI, multi-camera support, real-time monitoring, enhanced logging, etc.")

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
